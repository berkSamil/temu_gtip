"""
GTİP Sınıflandırma Değerlendirici
===================================
Gold set Excel'i alır, gtip_matcher ile sınıflandırır, 4 metrik hesaplar:
  - Fasıl    (ilk 2 hane)
  - Pozisyon (ilk 4 hane)
  - Alt poz  (ilk 6 hane)
  - Exact    (12 hane tam eşleşme)

Gold set format (data/gold_set.xlsx):
  Zorunlu : title, correct_gtip
  Opsiyonel: description, material, category, product_details, image_url

Kullanım:
    python scripts/eval_gtip.py data/gold_set.xlsx
    python scripts/eval_gtip.py data/gold_set.xlsx --db data/gtip_2026.db
    python scripts/eval_gtip.py data/gold_set.xlsx --refine --model claude-sonnet-4-20250514
    python scripts/eval_gtip.py data/gold_set.xlsx --out output/eval_result.xlsx
"""

import sys
import os
import re
import json
import sqlite3
import argparse
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("openpyxl yüklü değil: pip install openpyxl")
    sys.exit(1)

try:
    import anthropic
except ImportError:
    print("anthropic yüklü değil: pip install anthropic")
    sys.exit(1)

# gtip_matcher modülünü import et (aynı klasörde)
sys.path.insert(0, os.path.dirname(__file__))
from gtip_matcher import (
    classify_product,
    normalize_product_row,
    normalize_gtip_code,
)


# ---------------------------------------------------------------------------
# Gold set okuma
# ---------------------------------------------------------------------------

def read_gold_set(filepath):
    """
    Gold set Excel'i oku.
    correct_gtip kolonu zorunlu; title veya product_title zorunlu.
    Returns: list of dict
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sh = wb.active

    headers = []
    for cell in sh[1]:
        v = cell.value
        headers.append(str(v).strip() if v is not None else '')

    rows = []
    for r in sh.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in r):
            continue
        row = {headers[i]: (r[i] if r[i] is not None else '') for i in range(len(headers))}
        rows.append(row)

    if not rows:
        print("Hata: gold set boş")
        sys.exit(1)

    # correct_gtip kontrolü
    def normalize_col(s):
        return s.lower().strip().replace('i̇', 'i').replace('ı', 'i').replace('ğ', 'g').replace('ü', 'u').replace('ş', 's').replace('ö', 'o').replace('ç', 'c')

    cols_lower = {normalize_col(h): h for h in headers}
    gtip_col = None
    for candidate in ['correct_gtip', 'correct gtip', 'gtip', 'dogru_gtip', 'dogru gtip']:
        if candidate in cols_lower:
            gtip_col = cols_lower[candidate]
            break
    if gtip_col is None:
        print(f"Hata: 'correct_gtip' kolonu bulunamadı. Mevcut kolonlar: {headers}")
        sys.exit(1)

    out = []
    for i, row in enumerate(rows, 2):
        norm = normalize_product_row(row)
        raw_correct = str(row.get(gtip_col, '') or '').strip()
        correct = normalize_gtip_code(raw_correct)
        if not norm['title']:
            print(f"  Uyarı: Satır {i} — title boş, atlanıyor")
            continue
        if not correct:
            print(f"  Uyarı: Satır {i} — correct_gtip geçersiz ({raw_correct!r}), atlanıyor")
            continue
        norm['correct_gtip'] = correct
        out.append(norm)

    return out


# ---------------------------------------------------------------------------
# Metrik hesaplama
# ---------------------------------------------------------------------------

def _clean(code):
    """GTİP kodundan nokta kaldır: '3926.90.97.90.29' → '392690979029'"""
    return re.sub(r'[^0-9]', '', code or '')


def compute_metrics(correct, predicted):
    """
    correct, predicted: XXXX.XX.XX.XX.XX veya 4 haneli kod (veya boş)
    Returns: dict{fasil, pozisyon_secim, alt_poz, exact}  — her biri True/False/None
    None = predicted boş/geçersiz
    pozisyon_1b_kapsama eval döngüsünde ayrıca eklenir.
    """
    if not predicted:
        return {'fasil': None, 'pozisyon_secim': None, 'alt_poz': None, 'exact': None}

    c = _clean(correct)
    p = _clean(predicted)

    return {
        'fasil':           p[:2]  == c[:2],
        'pozisyon_secim':  p[:4]  == c[:4],
        'alt_poz':         p[:6]  == c[:6],
        'exact':           p[:12] == c[:12],
    }


def accuracy(hits, total, skipped):
    """hits/total yüzdesi; skipped = predicted boş olanlar."""
    if total == 0:
        return 0.0
    return hits / total * 100


# ---------------------------------------------------------------------------
# Excel çıktısı
# ---------------------------------------------------------------------------

def write_eval_excel(path, results):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Eval Sonuçları"

    headers = [
        'title', 'correct_gtip', 'predicted_gtip', 'guven',
        'fasil_ok', 'poz_secim_ok', 'poz_kapsama_ok', 'alt_poz_ok', 'exact_ok',
        'gerekce', 'soru', 'alternatifler', 'error'
    ]
    ws.append(headers)

    for r in results:
        m = r['metrics']
        def fmt(v):
            if v is None: return 'BILINEMEZ'
            return 'OK' if v else 'YANLIS'

        ws.append([
            r['title'],
            r['correct_gtip'],
            r['predicted_gtip'],
            r['guven'],
            fmt(m.get('fasil')),
            fmt(m.get('pozisyon_secim')),
            fmt(m.get('pozisyon_1b_kapsama')),
            fmt(m.get('alt_poz')),
            fmt(m.get('exact')),
            r['gerekce'][:300] if r['gerekce'] else '',
            r.get('soru', ''),
            ', '.join(r.get('alternatifler') or []),
            r.get('error', ''),
        ])

    from openpyxl.styles import PatternFill
    green = PatternFill(fill_type='solid', fgColor='C6EFCE')
    red   = PatternFill(fill_type='solid', fgColor='FFC7CE')
    grey  = PatternFill(fill_type='solid', fgColor='D9D9D9')

    col_map = {5: 'fasil', 6: 'pozisyon_secim', 7: 'pozisyon_1b_kapsama', 8: 'alt_poz', 9: 'exact'}
    for row_idx, r in enumerate(results, 2):
        m = r['metrics']
        for col_idx, metric_key in col_map.items():
            val = m.get(metric_key)
            cell = ws.cell(row=row_idx, column=col_idx)
            if val is None:
                cell.fill = grey
            elif val:
                cell.fill = green
            else:
                cell.fill = red

    wb.save(path)


# ---------------------------------------------------------------------------
# JSON experiment kaydı
# ---------------------------------------------------------------------------

_PROMPT_KEYS = {
    'bolum_system_prompt', 'bolum_user_msg',
    'fasil_system_prompt', 'fasil_user_msg',
    'pozisyon_system_prompt', 'pozisyon_context_block', 'pozisyon_query',
    'gtip_context_block', 'gtip_query',
}

def _build_result_entry(r, log_prompts=False):
    dbg = r.get('debug', {}) or {}
    if not log_prompts:
        dbg = {k: v for k, v in dbg.items() if k not in _PROMPT_KEYS}
    return {
        'title':          r['title'][:80],
        'correct_gtip':   r['correct_gtip'],
        'predicted_gtip': r['predicted_gtip'],
        'guven':          r['guven'],
        'soru':           r.get('soru', ''),
        'metrics':        r['metrics'],
        'debug':          dbg,
    }


def save_experiment(exp_dir, run_data):
    os.makedirs(exp_dir, exist_ok=True)
    ts = datetime.now().strftime('%Y%m%d_%H%M')
    path = os.path.join(exp_dir, f'run_{ts}.json')
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(run_data, f, ensure_ascii=False, indent=2)
    return path


# ---------------------------------------------------------------------------
# Paralel işleme
# ---------------------------------------------------------------------------

def _process_one(task):
    i, row, db_path, client, opts = task
    conn = sqlite3.connect(db_path)
    try:
        cls = classify_product(client, row, conn, opts)
        predicted = cls.get('gtip_code', '') or ''
        correct   = row['correct_gtip']
        metrics   = compute_metrics(correct, predicted)

        adim1b_parsed = cls.get('debug', {}).get('adim1b_parsed') or {}
        deger_dict    = adim1b_parsed.get('degerlendirme') or {}
        correct_poz4  = re.sub(r'[^0-9]', '', correct)[:4] if correct else ''
        if deger_dict and correct_poz4:
            metrics['pozisyon_1b_kapsama'] = any(
                re.sub(r'[^0-9]', '', str(k))[:4] == correct_poz4
                and isinstance(v, dict) and (v.get('karar') or '').strip() == 'Uyar'
                for k, v in deger_dict.items()
            )
        else:
            metrics['pozisyon_1b_kapsama'] = None

        def _poz_tanim(gtip_code):
            if not gtip_code:
                return ''
            poz4 = re.sub(r'[^0-9]', '', gtip_code)[:4]
            row_db = conn.execute(
                "SELECT tanim FROM pozisyon WHERE substr(kod_clean,1,4) = ? ORDER BY seviye LIMIT 1",
                (poz4,)
            ).fetchone()
            if row_db:
                return row_db[0]
            g = conn.execute(
                "SELECT tanim_hiyerarsi FROM gtip WHERE substr(gtip_clean,1,4) = ? ORDER BY gtip_code LIMIT 1",
                (poz4,)
            ).fetchone()
            return g[0] if g else ''

        dbg = cls.get('debug', {})
        if dbg is not None:
            dbg['correct_poz']       = re.sub(r'[^0-9]', '', correct)[:4]
            dbg['correct_poz_tanim'] = _poz_tanim(correct)
            dbg['pred_poz']          = re.sub(r'[^0-9]', '', predicted)[:4] if predicted else ''
            dbg['pred_poz_tanim']    = _poz_tanim(predicted)

        result = {
            'title':          row['title'],
            'correct_gtip':   correct,
            'predicted_gtip': predicted,
            'guven':          cls.get('guven', ''),
            'gerekce':        cls.get('gerekce', ''),
            'soru':           cls.get('soru', ''),
            'alternatifler':  cls.get('alternatifler', []),
            'error':          cls.get('error', ''),
            'metrics':        metrics,
            'debug':          dbg,
        }
        return i, result
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description='GTİP Sınıflandırma Değerlendirici',
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument('gold',            help='Gold set Excel yolu (correct_gtip kolonu zorunlu)')
    parser.add_argument('--db',            default='data/gtip_2026.db')
    parser.add_argument('--out',           default=None, help='Çıktı Excel yolu (varsayılan: output/eval_YYYYMMDD.xlsx)')
    parser.add_argument('--model',         default='claude-haiku-4-5-20251001')
    parser.add_argument('--max-tokens',    type=int, default=1200)
    parser.add_argument('--note-chars',    type=int, default=0)
    parser.add_argument('--gtip-rows',     type=int, default=120)
    parser.add_argument('--retrieval',     type=int, default=50)
    parser.add_argument('--delay',         type=float, default=0)
    parser.add_argument('--refine',        action='store_true')
    parser.add_argument('--refine-model',  default='claude-sonnet-4-20250514')
    parser.add_argument('--refine-max-tokens', type=int, default=1200)
    parser.add_argument('--experiments',   default='experiments', help='Experiment JSON klasörü')
    parser.add_argument('--limit',          type=int, default=None, help='İlk N ürünü çalıştır (test için)')
    parser.add_argument('--items',          default=None, help='Virgülle ayrılmış 1-tabanlı indeksler (orn. 6,21,25)')
    parser.add_argument('--izahname-chars', type=int, default=0, help='İzahname max karakter (0=kapalı)')
    parser.add_argument('--token-breakdown', action='store_true', help='Her atomun token sayısını JSON\'a yaz')
    parser.add_argument('--log-prompts',    action='store_true', help='Modele gönderilen tüm promptları JSON\'a yaz (dosya büyür)')
    parser.add_argument('--no-adim1b',      action='store_true', help='Adım 1b izahname doğrulama adımını atla')
    parser.add_argument('--adim1b-model',   default=None, help='Adım 1b için model (default: --model ile aynı)')
    parser.add_argument('--adim2',          action='store_true', help='Adım 2 GTİP seçimini etkinleştir (varsayılan: kapalı)')
    parser.add_argument('--provider',       default='deepseek', choices=['anthropic', 'deepseek'],
                        help='API provider (default: deepseek)')
    parser.add_argument('--workers',        type=int, default=50,
                        help='Paralel iş parçacığı sayısı (default: 50)')
    args = parser.parse_args()

    # .env yükle
    env_path = os.path.join(os.path.dirname(__file__), '..', '.env')
    if os.path.exists(env_path):
        with open(env_path) as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#') and '=' in line:
                    k, v = line.split('=', 1)
                    os.environ.setdefault(k.strip(), v.strip())

    _is_ds = args.provider == 'deepseek'
    if _is_ds:
        api_key = os.environ.get('DEEPSEEK_API_KEY', '')
        if not api_key:
            print("Hata: DEEPSEEK_API_KEY bulunamadı (.env veya ortam değişkeni)")
            sys.exit(1)
    else:
        api_key = os.environ.get('ANTHROPIC_API_KEY', '')
        if not api_key:
            print("Hata: ANTHROPIC_API_KEY bulunamadı (.env veya ortam değişkeni)")
            sys.exit(1)

    if not os.path.exists(args.db):
        print(f"Hata: DB bulunamadı: {args.db}")
        sys.exit(1)

    # Gold set oku
    print(f"Gold set yükleniyor: {args.gold}")
    gold_rows = read_gold_set(args.gold)
    if args.items:
        indices = [int(x.strip()) - 1 for x in args.items.split(',')]
        gold_rows = [gold_rows[i] for i in indices if 0 <= i < len(gold_rows)]
    elif args.limit:
        gold_rows = gold_rows[:args.limit]
    print(f"  {len(gold_rows)} ürün çalıştırılacak")
    print()

    conn = sqlite3.connect(args.db)

    if _is_ds:
        client = anthropic.Anthropic(
            base_url='https://api.deepseek.com/anthropic',
            api_key=api_key,
        )
    else:
        client = anthropic.Anthropic(api_key=api_key)

    opts = {
        'model':             args.model or ('deepseek-v4-flash' if _is_ds else 'claude-haiku-4-5-20251001'),
        'max_tokens':        args.max_tokens,
        'note_max_chars':    args.note_chars,
        'izahname_max_chars': args.izahname_chars,
        'token_breakdown':   args.token_breakdown,
        'gtip_rows_per_fasil': args.gtip_rows,
        'retrieval_top_n':   args.retrieval,
        'refine':            args.refine,
        'refine_model':      args.refine_model,
        'refine_max_tokens': args.refine_max_tokens,
        'adim1b':            not args.no_adim1b,
        'adim1b_model':      args.adim1b_model or ('deepseek-v4-pro' if _is_ds else 'claude-sonnet-4-20250514'),
        'adim2':             args.adim2,
        'provider':          args.provider,
    }

    n_items = len(gold_rows)
    tasks = [(i, row, args.db, client, opts) for i, row in enumerate(gold_rows, 1)]
    results_map = {}

    def _sym(v): return '✓' if v else ('?' if v is None else '✗')

    print(f"Paralel çalıştırılıyor ({min(args.workers, n_items)} worker)...\n")
    with ThreadPoolExecutor(max_workers=args.workers) as executor:
        futs = {executor.submit(_process_one, t): t[0] for t in tasks}
        for fut in as_completed(futs):
            i, result = fut.result()
            results_map[i] = result
            m = result['metrics']
            print(f"[{i:3d}/{n_items}] {(result['title'] or '')[:40]:40s}  "
                  f"→ {result['predicted_gtip'] or '(boş)':14s}  "
                  f"F:{_sym(m.get('fasil'))} P:{_sym(m.get('pozisyon_secim'))} "
                  f"K:{_sym(m.get('pozisyon_1b_kapsama'))}  [{result['guven']}]")

    results = [results_map[i] for i in range(1, n_items + 1)]

    # --- Özet hesapla ---
    n = len(results)
    skipped = sum(1 for r in results if not r['predicted_gtip'])

    def pct(key):
        hits  = sum(1 for r in results if r['metrics'].get(key) is True)
        denom = sum(1 for r in results if r['metrics'].get(key) is not None)
        return hits, denom, (hits / denom * 100 if denom else 0.0)

    f_h, f_d, f_acc = pct('fasil')
    p_h, p_d, p_acc = pct('pozisyon_secim')
    k_h, k_d, k_acc = pct('pozisyon_1b_kapsama')
    a_h, a_d, a_acc = pct('alt_poz')
    e_h, e_d, e_acc = pct('exact')

    print(f"\n{'='*62}")
    print(f"EVAL SONUÇLARI  ({n} ürün, {skipped} boş tahmin)")
    print(f"{'='*62}")
    print(f"  Fasıl          (2 hane) : {f_h:3d}/{f_d}  —  {f_acc:5.1f}%")
    print(f"  Poz Seçim      (4 hane) : {p_h:3d}/{p_d}  —  {p_acc:5.1f}%  ← pipeline tek seçim")
    print(f"  Poz Kapsama    (1b Uyar): {k_h:3d}/{k_d}  —  {k_acc:5.1f}%  ← Uyar listesinde var mı")
    if args.adim2:
        print(f"  Alt poz        (6 hane) : {a_h:3d}/{a_d}  —  {a_acc:5.1f}%")
        print(f"  Exact         (12 hane) : {e_h:3d}/{e_d}  —  {e_acc:5.1f}%")
    print(f"{'='*62}")

    # Fasıl bazında breakdown
    fasil_stats = {}
    for r in results:
        correct_fasil = _clean(r['correct_gtip'])[:2]
        if correct_fasil not in fasil_stats:
            fasil_stats[correct_fasil] = {'total': 0, 'fasil_ok': 0, 'exact_ok': 0}
        fasil_stats[correct_fasil]['total'] += 1
        if r['metrics'].get('fasil') is True:
            fasil_stats[correct_fasil]['fasil_ok'] += 1
        if r['metrics'].get('exact') is True:
            fasil_stats[correct_fasil]['exact_ok'] += 1

    if len(fasil_stats) > 1:
        print(f"\nFasıl bazında breakdown:")
        for fn, s in sorted(fasil_stats.items()):
            print(f"  Fasıl {fn}: {s['total']} ürün | "
                  f"fasıl %{s['fasil_ok']/s['total']*100:.0f} | "
                  f"exact %{s['exact_ok']/s['total']*100:.0f}")

    # --- Excel çıktısı ---
    if args.out is None:
        os.makedirs('output', exist_ok=True)
        ts = datetime.now().strftime('%Y%m%d_%H%M')
        args.out = f'output/eval_{ts}.xlsx'

    write_eval_excel(args.out, results)
    print(f"\nExcel kaydedildi : {args.out}")

    # --- Experiment JSON ---
    run_data = {
        'timestamp':    datetime.now().isoformat(),
        'provider':     args.provider,
        'model':        opts['model'],
        'adim1b_model': opts['adim1b_model'],
        'refine':       args.refine,
        'refine_model': args.refine_model if args.refine else None,
        'n_total':    n,
        'n_skipped':  skipped,
        'metrics': {
            'fasil':               {'hits': f_h, 'total': f_d, 'accuracy': round(f_acc, 2)},
            'pozisyon_secim':      {'hits': p_h, 'total': p_d, 'accuracy': round(p_acc, 2)},
            'pozisyon_1b_kapsama': {'hits': k_h, 'total': k_d, 'accuracy': round(k_acc, 2)},
            'alt_poz':             {'hits': a_h, 'total': a_d, 'accuracy': round(a_acc, 2)},
            'exact':               {'hits': e_h, 'total': e_d, 'accuracy': round(e_acc, 2)},
        },
        'params': {
            'note_chars':  args.note_chars,
            'gtip_rows':   args.gtip_rows,
            'retrieval':   args.retrieval,
            'max_tokens':  args.max_tokens,
        },
        'results': [
            _build_result_entry(r, log_prompts=args.log_prompts)
            for r in results
        ]
    }

    exp_path = save_experiment(args.experiments, run_data)
    print(f"Experiment JSON  : {exp_path}")


if __name__ == '__main__':
    main()
