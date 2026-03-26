"""
TEMU Urun Bilgisi Scraper
==========================
Excel'deki TEMU urun linklerinden urun bilgilerini ceker.
Chrome Debug modunda acilmis tarayiciya baglanir (CDP).

Kullanim:
  1) Masaustundeki 'Chrome Debug.bat' calistirin
  2) Temu hesabiniza giris yapin (ilk seferde)
  3) Scripti calistirin:
       python temu_scraper.py data/input.xlsx
       python temu_scraper.py data/input.xlsx --delay 12 --jitter 8
       python temu_scraper.py data/input.xlsx --warmup-every 12
       python temu_scraper.py data/input.xlsx --skip 0 --limit 20 -o output/part1.xlsx
       python temu_scraper.py data/input.xlsx --skip 20 --limit 20 -o output/part2.xlsx
       (Yavas = daha az "sold out" yaniltmasi; --warmup-every ~12-15 sik softblock icin.)

  Ipuclari: Chrome Debug profilinde mumkunse giris acik tutun; US IP + ulke uyumu;
  cok hizli ardisik goods.html yuklemesi riskli — gecikmeyi artirin.
"""

import sys
import os
import re
import json
import time
import random
import argparse
import urllib.parse
import urllib.request

try:
    import openpyxl
except ImportError:
    print("openpyxl yuklu degil: pip install openpyxl")
    sys.exit(1)

try:
    from playwright.sync_api import sync_playwright
except ImportError:
    print("playwright yuklu degil: pip install playwright && playwright install chromium")
    sys.exit(1)

DETAIL_WAIT_MS = 18000
PAGE_TIMEOUT_MS = 45000
CAPTCHA_TIMEOUT = 120
DEFAULT_WARMUP_URL = 'https://www.temu.com'


def detect_captcha(page):
    """Detect Temu security challenge (English UI even when region changes)."""
    return page.evaluate('''() => {
        const t = (document.body?.innerText || '').toLowerCase();
        if (t.includes('security verification')) return true;
        const el = document.querySelector('[class*="captcha" i], [class*="verify-wrap"], [id*="captcha" i]');
        return !!el;
    }''')


def extract_goods_id(url):
    m = re.search(r'goods_id=(\d+)', url) or re.search(r'-g-(\d+)\.html', url)
    return m.group(1) if m else ''


def extract_gallery_url(url):
    try:
        qs = urllib.parse.parse_qs(urllib.parse.urlparse(url).query)
        return qs.get('top_gallery_url', [''])[0]
    except Exception:
        return ''


def wait_for_captcha(page, timeout=CAPTCHA_TIMEOUT):
    """Detect CAPTCHA and wait for user to solve it."""
    if not detect_captcha(page):
        return False

    print("\n  *** CAPTCHA! Chrome penceresinde cozun... ***", end="", flush=True)
    for _ in range(timeout // 2):
        time.sleep(2)
        if not detect_captcha(page):
            print(" OK", flush=True)
            time.sleep(2)
            return True
    print(" TIMEOUT", flush=True)
    return True


def wait_for_product_data(page, timeout_ms=DETAIL_WAIT_MS):
    try:
        page.wait_for_function(
            '() => { try { return window.rawData.store.goodsProperty.length > 0 } catch(e) { return false } }',
            timeout=timeout_ms
        )
        return True
    except Exception:
        return False


def extract_product_data(page):
    result = page.evaluate('''() => {
        const s = window.rawData?.store || {};
        const g = s.goods || {};

        const props = (s.goodsProperty || []).map(p => ({
            key: p.key || '',
            values: (p.values || []).join(', ')
        }));

        const gallery = (g.gallery || []).map(img =>
            typeof img === 'object' ? (img.url || '') : String(img)
        ).filter(Boolean);

        const banner = (g.bannerList || []).map(b =>
            typeof b === 'object' ? (b.url || '') : String(b)
        ).filter(Boolean);

        const fsd = s.formatSkuData || {};
        const variants = (fsd.skuTypeValues || []).map(v => ({
            key: v.type || '', values: (v.values || []).join(', ')
        }));

        return {
            goodsName: g.goodsName || null,
            status: g.status ?? null,
            goodsProperty: props,
            variants: variants,
            galleryUrls: gallery.slice(0, 5),
            bannerUrls: banner.slice(0, 3),
            hdThumbUrl: g.hdThumbUrl || '',
        };
    }''')

    meta = page.evaluate('''() => {
        const get = (sel) => {
            const el = document.querySelector(sel);
            return el ? el.getAttribute('content') || '' : '';
        };
        return {
            title: document.title || '',
            description: get('meta[name="description"]'),
            keywords: get('meta[name="keywords"]'),
        };
    }''')

    return result, meta


def _raw_signal(page):
    """Status + goodsProperty count (Temu status 5 = sold out / unavailable in session)."""
    return page.evaluate('''() => {
        try {
            const s = window.rawData?.store;
            if (!s) return { status: null, gp: 0, hasRaw: false };
            const st = s.goods?.status;
            const gp = (s.goodsProperty || []).length;
            return { status: st, gp, hasRaw: true };
        } catch(e) { return { status: null, gp: 0, hasRaw: false }; }
    }''')


def human_activity(page):
    """Light activity so Temu sees less like a bare script."""
    try:
        page.mouse.move(random.randint(80, 280), random.randint(80, 280))
        time.sleep(random.uniform(0.25, 0.6))
        page.evaluate('window.scrollTo(0, 200 + Math.floor(Math.random()*200))')
        time.sleep(random.uniform(0.3, 0.7))
    except Exception:
        pass


def session_warmup(page, url=None):
    """
    Once oturumda Temu ana sayfasina git: soguk goods_id linkleri yerine
    once first-party sayfa (elle paste oncesi gezinme davranisina yakin).
    """
    target = (url or DEFAULT_WARMUP_URL).strip()
    if not target:
        return
    try:
        page.goto(target, wait_until='load', timeout=PAGE_TIMEOUT_MS)
        time.sleep(random.uniform(2.0, 5.0))
        wait_for_captcha(page)
        try:
            page.wait_for_load_state('networkidle', timeout=12000)
        except Exception:
            pass
        human_activity(page)
        time.sleep(random.uniform(1.0, 2.5))
    except Exception:
        pass


def safe_goto(page, url, retries=3):
    """Navigate with retry on context-destroyed errors (redirects)."""
    for attempt in range(retries):
        try:
            page.goto(url, wait_until='load', timeout=PAGE_TIMEOUT_MS)
            time.sleep(random.uniform(1.8, 3.2))
            wait_for_captcha(page)
            try:
                page.wait_for_load_state('networkidle', timeout=10000)
            except Exception:
                pass
            human_activity(page)
            return True
        except Exception as e:
            if 'context was destroyed' in str(e) or 'navigation' in str(e).lower():
                time.sleep(3)
                try:
                    page.wait_for_load_state('load', timeout=20000)
                    time.sleep(2)
                    wait_for_captcha(page)
                    human_activity(page)
                    return True
                except Exception:
                    if attempt < retries - 1:
                        continue
            raise
    return False


def scrape_product(page, url, delay=2.0):
    goods_id = extract_goods_id(url)
    gallery_from_url = extract_gallery_url(url)

    row = {
        'url': url,
        'goods_id': goods_id,
        'title': '',
        'description': '',
        'keywords': '',
        'product_details': '',
        'image_url': gallery_from_url,
        'properties': '',
        'error': '',
    }

    try:
        safe_goto(page, url)

        loaded = wait_for_product_data(page)

        if not loaded:
            if detect_captcha(page):
                wait_for_captcha(page)
                loaded = wait_for_product_data(page)
            else:
                page.evaluate('window.scrollTo(0, 500)')
                time.sleep(5)
                human_activity(page)
                loaded = wait_for_product_data(page, timeout_ms=12000)

        sig = _raw_signal(page)
        if sig.get('status') == 5 or (sig.get('hasRaw') and sig.get('gp', 0) == 0):
            time.sleep(random.uniform(4, 8))
            wait_for_captcha(page)
            try:
                page.reload(wait_until='load', timeout=PAGE_TIMEOUT_MS)
            except Exception:
                pass
            time.sleep(random.uniform(2.5, 4.5))
            wait_for_captcha(page)
            human_activity(page)
            loaded = wait_for_product_data(page, timeout_ms=DETAIL_WAIT_MS)

        sig = _raw_signal(page)
        if not loaded or sig.get('gp', 0) == 0:
            time.sleep(random.uniform(6, 12))
            wait_for_captcha(page)
            safe_goto(page, url, retries=2)
            loaded = wait_for_product_data(page, timeout_ms=DETAIL_WAIT_MS)

        product, meta = extract_product_data(page)

        name = product.get('goodsName') or ''
        if not name or name == 'Unavailable for purchase':
            name = meta.get('title', '')
            name = re.sub(r'\s*[-\u2013]\s*Temu\b.*$', '', name).strip()
            if name.lower() in ('temu', ''):
                slug_m = re.search(r'temu\.com/(?:[a-z]{2}(?:-[a-z]{2})?/)?(.+?)(?:-g-\d+\.html|$)', url)
                if slug_m:
                    name = slug_m.group(1).replace('-', ' ').strip()
        row['title'] = name

        desc = meta.get('description', '')
        desc = re.sub(r'^(Shop |Check out this |Find |Free returns\.\s*)', '', desc).strip()
        row['description'] = desc
        row['keywords'] = meta.get('keywords', '').replace('&amp;', '&')

        gp = product.get('goodsProperty', [])
        if gp:
            details = [f"{p['key']}: {p['values']}" for p in gp if p.get('key') and p.get('values')]
            row['product_details'] = '; '.join(details)

        variants = product.get('variants', [])
        if variants:
            row['properties'] = '; '.join(f"{v['key']}: {v['values']}" for v in variants if v.get('key'))

        imgs = product.get('galleryUrls', [])
        if imgs:
            row['image_url'] = imgs[0]
        elif product.get('bannerUrls'):
            row['image_url'] = product['bannerUrls'][0]
        elif product.get('hdThumbUrl'):
            row['image_url'] = product['hdThumbUrl']

    except Exception as e:
        row['error'] = str(e)[:100]

    return row


def read_links(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    links = []
    for row in ws.iter_rows(min_row=1, max_col=ws.max_column, values_only=False):
        for cell in row:
            val = str(cell.value or '').strip()
            if 'temu.com' in val and ('goods' in val or '-g-' in val):
                links.append(val)
    return links


def write_output(results, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TEMU Products"

    headers = ['URL', 'Goods ID', 'Title', 'Description', 'Keywords',
               'Product Details', 'Image URL', 'SKU Variants', 'Error']
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        ws.cell(1, col).font = openpyxl.styles.Font(bold=True)

    for r in results:
        ws.append([
            r['url'], r['goods_id'], r['title'], r['description'],
            r['keywords'], r['product_details'], r['image_url'],
            r['properties'], r['error']
        ])

    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 60
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 60

    wb.save(output_path)


def _esc(text):
    """Escape HTML special characters."""
    return (text or '').replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;').replace('"', '&quot;')


def write_html(results, output_path):
    html_path = os.path.splitext(output_path)[0] + '.html'

    cards = []
    for i, r in enumerate(results, 1):
        details_html = ''
        if r['product_details']:
            rows = ''
            for item in r['product_details'].split('; '):
                parts = item.split(': ', 1)
                if len(parts) == 2:
                    rows += f'<tr><td class="prop-key">{_esc(parts[0])}</td><td>{_esc(parts[1])}</td></tr>'
            details_html = f'<table class="props">{rows}</table>'

        variants_html = ''
        if r['properties']:
            for v in r['properties'].split('; '):
                variants_html += f'<span class="variant">{_esc(v)}</span> '

        img_html = ''
        if r['image_url']:
            img_html = f'<img src="{_esc(r["image_url"])}" alt="{_esc(r["title"])}" loading="lazy">'

        error_html = ''
        if r['error']:
            error_html = f'<div class="error">{_esc(r["error"])}</div>'

        cards.append(f'''
    <div class="card">
      <div class="card-img">{img_html}</div>
      <div class="card-body">
        <div class="card-num">#{i}</div>
        <h2><a href="{_esc(r['url'])}" target="_blank">{_esc(r['title']) or 'Untitled'}</a></h2>
        <p class="desc">{_esc(r['description'])}</p>
        {details_html}
        {f'<div class="variants">{variants_html}</div>' if variants_html else ''}
        <div class="meta">
          <span>ID: {_esc(r['goods_id'])}</span>
          {f'<span class="kw">{_esc(r["keywords"][:80])}</span>' if r['keywords'] else ''}
        </div>
        {error_html}
      </div>
    </div>''')

    ok = sum(1 for r in results if not r['error'])
    det = sum(1 for r in results if r['product_details'])

    html = f'''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>TEMU Scrape Results ({len(results)} products)</title>
<style>
  * {{ margin: 0; padding: 0; box-sizing: border-box; }}
  body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; background: #f0f2f5; color: #1a1a1a; padding: 20px; }}
  .header {{ max-width: 1100px; margin: 0 auto 24px; }}
  .header h1 {{ font-size: 24px; font-weight: 700; }}
  .header .stats {{ color: #666; margin-top: 4px; font-size: 14px; }}
  .stats span {{ margin-right: 16px; }}
  .card {{ max-width: 1100px; margin: 0 auto 16px; background: #fff; border-radius: 10px; box-shadow: 0 1px 3px rgba(0,0,0,.08); display: flex; overflow: hidden; }}
  .card-img {{ width: 220px; min-height: 220px; flex-shrink: 0; background: #f7f7f7; display: flex; align-items: center; justify-content: center; }}
  .card-img img {{ width: 100%; height: 100%; object-fit: cover; }}
  .card-body {{ padding: 16px 20px; flex: 1; min-width: 0; }}
  .card-num {{ font-size: 12px; color: #999; margin-bottom: 4px; }}
  h2 {{ font-size: 16px; font-weight: 600; margin-bottom: 8px; line-height: 1.3; }}
  h2 a {{ color: #1a1a1a; text-decoration: none; }}
  h2 a:hover {{ color: #e67e00; }}
  .desc {{ font-size: 13px; color: #555; margin-bottom: 10px; line-height: 1.4; display: -webkit-box; -webkit-line-clamp: 2; -webkit-box-orient: vertical; overflow: hidden; }}
  .props {{ font-size: 13px; border-collapse: collapse; margin-bottom: 10px; }}
  .props tr {{ border-bottom: 1px solid #f0f0f0; }}
  .props td {{ padding: 3px 12px 3px 0; }}
  .prop-key {{ font-weight: 600; color: #333; white-space: nowrap; }}
  .variants {{ margin-bottom: 8px; }}
  .variant {{ display: inline-block; background: #f0f2f5; padding: 2px 8px; border-radius: 4px; font-size: 12px; margin: 2px 4px 2px 0; }}
  .meta {{ font-size: 11px; color: #999; }}
  .meta span {{ margin-right: 12px; }}
  .kw {{ max-width: 300px; display: inline-block; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; vertical-align: bottom; }}
  .error {{ margin-top: 8px; padding: 6px 10px; background: #fff0f0; color: #c00; border-radius: 4px; font-size: 13px; }}
  @media (max-width: 700px) {{
    .card {{ flex-direction: column; }}
    .card-img {{ width: 100%; height: 200px; }}
  }}
</style>
</head>
<body>
<div class="header">
  <h1>TEMU Scrape Results</h1>
  <div class="stats">
    <span>{len(results)} products</span>
    <span>{ok} successful</span>
    <span>{det} with details</span>
  </div>
</div>
{"".join(cards)}
</body>
</html>'''

    with open(html_path, 'w', encoding='utf-8') as f:
        f.write(html)

    return html_path


def download_image(image_url, save_dir, goods_id):
    if not image_url:
        return ''
    try:
        ext = '.jpg'
        if '.png' in image_url:
            ext = '.png'
        elif '.webp' in image_url:
            ext = '.webp'
        fpath = os.path.join(save_dir, f"{goods_id}{ext}")
        req = urllib.request.Request(image_url, headers={
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)'
        })
        resp = urllib.request.urlopen(req, timeout=15)
        with open(fpath, 'wb') as f:
            f.write(resp.read())
        return fpath
    except Exception:
        return ''


def main():
    parser = argparse.ArgumentParser(description='TEMU Product Scraper')
    parser.add_argument('input', help='Excel file with TEMU links')
    parser.add_argument('-o', '--output', help='Output Excel file')
    parser.add_argument('--images', action='store_true', help='Download product images')
    parser.add_argument('--delay', type=float, default=8.0,
                        help='Base delay between products (seconds); Temu throttles fast runs')
    parser.add_argument('--jitter', type=float, default=5.0,
                        help='Random extra 0..jitter seconds added after each product')
    parser.add_argument('--port', type=int, default=9222, help='Chrome debugging port')
    parser.add_argument(
        '--no-warmup',
        action='store_true',
        help='Temu ana sayfa oturum isitmasini atla',
    )
    parser.add_argument(
        '--warmup-url',
        default=DEFAULT_WARMUP_URL,
        help='Oturum isitma icin acilacak URL (varsayilan: https://www.temu.com)',
    )
    parser.add_argument(
        '--warmup-every',
        type=int,
        default=0,
        metavar='N',
        help='Her N urunden sonra tekrar isit (0=sadece baslangicta). Or: 12-15 softblock icin',
    )
    parser.add_argument(
        '--skip',
        type=int,
        default=0,
        metavar='K',
        help='Exceldeki link listesinde ilk K linki atla (20ser parti: 0, 20, 40, ...)',
    )
    parser.add_argument(
        '--limit',
        type=int,
        default=None,
        metavar='N',
        help='En fazla N link isle (20ser parti icin: 20). Verilmezse skip sonrasi hepsi',
    )
    args = parser.parse_args()

    if not os.path.isfile(args.input):
        print(f"Hata: {args.input} bulunamadi")
        sys.exit(1)

    output_path = args.output
    base = os.path.splitext(os.path.basename(args.input))[0]
    if not output_path:
        if args.skip or args.limit is not None:
            lim_part = args.limit if args.limit is not None else 'all'
            output_path = os.path.join(
                'output', f'{base}_skip{args.skip}_lim{lim_part}_scraped.xlsx'
            )
        else:
            output_path = os.path.join('output', f'{base}_scraped.xlsx')
    os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)

    image_dir = None
    if args.images:
        image_dir = os.path.join(os.path.dirname(output_path), 'images')
        os.makedirs(image_dir, exist_ok=True)

    all_links = read_links(args.input)
    skip = max(0, args.skip)
    if args.limit is not None:
        links = all_links[skip : skip + max(0, args.limit)]
    else:
        links = all_links[skip:]

    print(f"Excelde toplam link: {len(all_links)} | Bu kosu: {len(links)} (skip={skip})")
    if not all_links:
        print("Hata: Excel'de TEMU linki bulunamadi")
        sys.exit(1)
    if not links:
        print("Hata: --skip / --limit sonrasi islenecek link kalmadi")
        sys.exit(1)

    print(f"Chrome'a baglaniliyor (port {args.port})...", flush=True)

    with sync_playwright() as pw:
        try:
            browser = pw.chromium.connect_over_cdp(f'http://127.0.0.1:{args.port}')
        except Exception:
            print(
                "\nChrome'a baglanilamadi!\n"
                "Masaustundeki 'Chrome Debug.bat' dosyasini cift tiklayin,\n"
                "Chrome acildiktan sonra scripti tekrar calistirin.\n"
            )
            sys.exit(1)

        ctx = browser.contexts[0]
        page = ctx.new_page()
        print("Baglandi.\n")

        if not args.no_warmup:
            print("Oturum isitma (Temu)...", flush=True)
            session_warmup(page, args.warmup_url)
            print("", flush=True)

        results = []
        errors = 0

        for i, url in enumerate(links, 1):
            if (
                not args.no_warmup
                and args.warmup_every > 0
                and i > 1
                and (i - 1) % args.warmup_every == 0
            ):
                print(f"  [oturum isitma #{i}]...", end=" ", flush=True)
                session_warmup(page, args.warmup_url)

            gid = extract_goods_id(url)
            print(f"  [{i}/{len(links)}] {gid or url[:50]}...", end=" ", flush=True)

            row = scrape_product(page, url, args.delay)

            if row['error']:
                print(f"HATA: {row['error']}")
                errors += 1
            else:
                dp = len(row['product_details'].split(';')) if row['product_details'] else 0
                tag = f"[{dp} details]" if dp > 0 else "[meta only]"
                print(f"{row['title'][:50]} {tag}")

            results.append(row)

            if i < len(links):
                pause = args.delay + random.uniform(0, max(0.0, args.jitter))
                time.sleep(pause)

        page.close()
        browser.close()

    if args.images and image_dir:
        print("\nResimler indiriliyor...")
        for r in results:
            if r['image_url'] and r['goods_id']:
                download_image(r['image_url'], image_dir, r['goods_id'])

    write_output(results, output_path)
    html_path = write_html(results, output_path)

    ok = len(results) - errors
    details = sum(1 for r in results if r['product_details'])
    imgs = sum(1 for r in results if r['image_url'])
    print(f"\nToplam         : {len(results)}")
    print(f"Basarili       : {ok}")
    print(f"Hata           : {errors}")
    print(f"Product Details: {details}/{ok}")
    print(f"Resim          : {imgs}/{ok}")
    print(f"Excel          : {output_path}")
    print(f"HTML           : {html_path}")


if __name__ == "__main__":
    main()
