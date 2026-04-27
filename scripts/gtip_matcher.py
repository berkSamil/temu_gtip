"""
GTIP Siniflandirici
=====================
TEMU scraper ciktisi veya manuel doldurulmus Excel (product url, product title,
product details, ...) ile GTIP siniflandirma; Claude + SQLite cetvel.

Kullanim:
    python gtip_matcher.py output/input_scraped.xlsx
    python gtip_matcher.py output/input_scraped.xlsx --db data/gtip_2026.db -o output/classified.xlsx
    python gtip_matcher.py input.xlsx --refine --note-chars 6000
    python gtip_matcher.py input.xlsx --model claude-sonnet-4-20250514 --max-tokens 1600
"""

import sys
import os
import re
import json
import sqlite3
import argparse
import time

try:
    import openpyxl
except ImportError:
    print("openpyxl yuklu degil: pip install openpyxl")
    sys.exit(1)

try:
    import anthropic
except ImportError:
    print("anthropic yuklu degil: pip install anthropic")
    sys.exit(1)


# ---------------------------------------------------------------------------
# DB queries
# ---------------------------------------------------------------------------

_TEMU_STOP = frozenset({
    'the', 'and', 'for', 'with', 'from', 'this', 'that', 'your', 'are', 'you', 'all', 'any',
    'can', 'has', 'have', 'pcs', 'pack', 'set', 'piece', 'pieces', 'item', 'items', 'sale',
    'shop', 'temu', 'free', 'new', 'hot', 'best', 'buy', 'get', 'one', 'two', 'off', 'out',
    'our', 'was', 'not', 'but', 'its', 'per', 'use', 'may', 'more', 'most', 'some', 'size',
})


def get_candidate_fasils(conn, product_details, keywords, description, title, max_fasils=8):
    """
    Aday fasillari belirle: urun metninden kelimeler -> FTS (gtip_fts + notlar_fts) ile
    eslesen satirlarin fasil numaralari. Material hints kaldirildi — tarife mantığına aykırı.
    """
    text = f"{title} {description} {keywords} {product_details}".lower()
    scores = {}

    words = sorted(
        set(re.findall(r'[a-zA-ZğüşıöçĞÜŞİÖÇ]{3,}', text)),
        key=len,
        reverse=True,
    )
    words = [w for w in words if w.lower() not in _TEMU_STOP][:14]

    for w in words:
        rows = search_gtip_fts(conn, w, limit=18)
        for r in rows:
            code = r[0]
            parts = str(code).split('.')
            if not parts or not parts[0].isdigit():
                continue
            fn = int(parts[0][:2])
            scores[fn] = scores.get(fn, 0) + 1

    ordered = sorted(scores.keys(), key=lambda x: (-scores[x], x))
    out = []
    seen = set()
    for fn in ordered:
        if fn not in seen:
            seen.add(fn)
            out.append(fn)
        if len(out) >= max_fasils:
            return out

    defaults = [39, 73, 82, 83, 84, 85, 90, 94, 96, 61, 62, 33, 42, 95, 87, 71, 91, 48, 64]
    for d in defaults:
        if d not in seen:
            seen.add(d)
            out.append(d)
        if len(out) >= max_fasils:
            break
    return out


def get_fasil_gtip_list(conn, fasil_no, limit=200):
    c = conn.cursor()
    rows = c.execute("""
        SELECT gtip_code, tanim, tanim_hiyerarsi
        FROM gtip WHERE fasil = ?
        ORDER BY gtip_code
        LIMIT ?
    """, (fasil_no, limit)).fetchall()
    return rows


def get_fasil_notu(conn, fasil_no):
    c = conn.cursor()
    row = c.execute(
        "SELECT fasil_notu FROM fasil_notlari WHERE fasil_no = ?",
        (fasil_no,)
    ).fetchone()
    return row[0] if row else ""


def get_all_pozisyonlar(conn, fasil_no):
    """
    Fasılın tüm 4'lü pozisyonlarını döner.

    Tarife cetveli yarı-düzenlidir: bazı pozisyonlar cetvelde 4'lü değil
    doğrudan 6'lı seviyede başlar (örn. 9602.00, 9605.00). Bu durumda
    pozisyon tablosunda seviye=4 kaydı olmaz ama gtip tablosunda ilgili
    12'liler mevcuttur. Yalnızca pozisyon tablosuna bağlı kalmak bu
    pozisyonları Step 1'den gizler ve Step 2'ye geçilmesini engeller.

    Çözüm: gtip tablosundan substr(gtip_clean,1,4) ile sentetik 4'lü
    gruplar türet, açıklama için pozisyon tablosuna bak, yoksa gtip
    hiyerarşi metnini kullan.
    """
    c = conn.cursor()
    # 1) Fasıldaki tüm benzersiz 4-hane öneklerini GTİP tablosundan çıkar
    prefixes = [
        row[0]
        for row in c.execute("""
            SELECT DISTINCT substr(gtip_clean, 1, 4) AS p
            FROM gtip
            WHERE fasil = ?
            ORDER BY p
        """, (fasil_no,)).fetchall()
    ]
    result = []
    for pref in prefixes:
        # 2a) pozisyon tablosunda bu önekle başlayan en kısa kaydı ara
        row = c.execute("""
            SELECT tanim FROM pozisyon
            WHERE fasil = ? AND substr(kod_clean, 1, 4) = ?
            ORDER BY seviye
            LIMIT 1
        """, (fasil_no, pref)).fetchone()
        if row:
            result.append((pref, row[0]))
        else:
            # gtip tablosundan tanim_hiyerarsi ile temsil et
            g = c.execute("""
                SELECT tanim_hiyerarsi FROM gtip
                WHERE fasil = ? AND substr(gtip_clean, 1, 4) = ?
                ORDER BY gtip_code
                LIMIT 1
            """, (fasil_no, pref)).fetchone()
            if g:
                result.append((pref, g[0] or pref))
    return result


def get_gtips_by_pozisyon(conn, pozisyon_kod):
    """Seçilen 4'lü pozisyon altındaki tüm 12'li GTİP'leri döner."""
    clean = re.sub(r'[^0-9]', '', str(pozisyon_kod))[:4]
    c = conn.cursor()
    return c.execute("""
        SELECT gtip_code, tanim, tanim_hiyerarsi FROM gtip
        WHERE gtip_clean LIKE ?
        ORDER BY gtip_code
    """, (clean + '%',)).fetchall()


def get_izahname(conn, fasil_no, max_chars=3000):
    """Fasıl izahname metnini döner (kırpılmış)."""
    c = conn.cursor()
    row = c.execute(
        "SELECT metin FROM izahname_notlari WHERE fasil_no = ?",
        (fasil_no,)
    ).fetchone()
    if not row or not row[0]:
        return ""
    return row[0][:max_chars]


def get_izahname_for_pozisyon(conn, fasil_no, poz4):
    """Fasıl izahname metninden belirtilen 4'lü pozisyona ait bölümü tam olarak çıkarır.

    poz4: '3924', '9603' gibi 4 haneli kod (noktasız).
    Sonraki pozisyon başlığına kadar tüm metni döner (kesmez).
    """
    row = conn.execute(
        "SELECT metin FROM izahname_notlari WHERE fasil_no = ?", (fasil_no,)
    ).fetchone()
    if not row or not row[0]:
        return ""
    text = row[0]
    poz_dotted = f"{poz4[:2]}.{poz4[2:]}"
    # Satır başından başlayan bölüm başlığını bul (inline referansları atla)
    m = re.search(r'(?:^|\n)' + re.escape(poz_dotted) + r'[\s\-\t]', text)
    if not m:
        return ""
    start = m.start()
    # Satır başı (\n) eşleşmesinde asıl içerik bir sonraki karakterden başlar
    if text[start] == '\n':
        start += 1
    # Sonraki pozisyon başlığına kadar al (satır başında XX.XX formatı)
    m2 = re.search(r'\n\d{2}\.\d{2}[\s\-\t]', text[start + 1:])
    end = (start + 1 + m2.start()) if m2 else len(text)
    return text[start:end].strip()


def get_yorum_kurallari(conn):
    """Tüm yorum kurallarını tek metin olarak döner (özet)."""
    c = conn.cursor()
    rows = c.execute(
        "SELECT kural_no, metin FROM yorum_kurallari ORDER BY kural_no"
    ).fetchall()
    parts = []
    for kural_no, metin in rows:
        parts.append(f"KURAL {kural_no}:\n{(metin or '')[:600]}")
    return "\n\n".join(parts)


_TAXONOMY_CACHE = {}
_PROMPT_CACHE = {}
_BOLUM_CACHE = {}


def get_bolum_listesi(conn):
    """21 bölümü döner: [(bolum_no, bolum_adi), ...]. Sonuç cache'lenir."""
    if 'bolumler' in _BOLUM_CACHE:
        return _BOLUM_CACHE['bolumler']
    c = conn.cursor()
    rows = c.execute(
        "SELECT DISTINCT bolum_no, bolum_adi FROM bolum_fasil ORDER BY bolum_no"
    ).fetchall()
    _BOLUM_CACHE['bolumler'] = rows
    return rows


def get_fasiller_by_bolumler(conn, bolum_nos):
    """Verilen bölüm numaralarının fasıllarını döner: [(fasil_no, fasil_adi), ...]."""
    if not bolum_nos:
        return []
    if 'fasil_map' not in _BOLUM_CACHE:
        c = conn.cursor()
        _BOLUM_CACHE['fasil_map'] = c.execute(
            "SELECT bolum_no, fasil_no, fasil_adi FROM bolum_fasil ORDER BY bolum_no, fasil_no"
        ).fetchall()
    bolum_set = set(bolum_nos)
    return [(fasil_no, fasil_adi)
            for bolum_no, fasil_no, fasil_adi in _BOLUM_CACHE['fasil_map']
            if bolum_no in bolum_set]


def get_fasil_taxonomy(conn, note_chars=300):
    """
    bolum_fasil tablosundan tüm bölüm→fasıl hiyerarşisini döner.
    Her fasıl için kısa fasıl notu özeti de eklenir. Sonuç cache'lenir.
    """
    cache_key = note_chars
    if cache_key in _TAXONOMY_CACHE:
        return _TAXONOMY_CACHE[cache_key]

    c = conn.cursor()
    rows = c.execute(
        "SELECT bolum_no, bolum_adi, fasil_no, fasil_adi FROM bolum_fasil ORDER BY bolum_no, fasil_no"
    ).fetchall()

    parts = []
    current_bolum = None
    for bolum_no, bolum_adi, fasil_no, fasil_adi in rows:
        if bolum_no != current_bolum:
            parts.append(f"\nBÖLÜM {bolum_no}: {bolum_adi}")
            current_bolum = bolum_no
        if note_chars > 0:
            note = get_fasil_notu(conn, fasil_no)
            note_excerpt = f" | {note[:note_chars]}" if note else ""
        else:
            note_excerpt = ""
        parts.append(f"  Fasıl {fasil_no:02d}: {fasil_adi}{note_excerpt}")

    result = "\n".join(parts)
    _TAXONOMY_CACHE[cache_key] = result
    return result


def build_bolum_prompt():
    if 'bolum' in _PROMPT_CACHE:
        return _PROMPT_CACHE['bolum']
    _PROMPT_CACHE['bolum'] = _BOLUM_PROMPT_BASE
    return _BOLUM_PROMPT_BASE


def build_fasil_prompt(conn):
    if 'fasil' in _PROMPT_CACHE:
        return _PROMPT_CACHE['fasil']
    result = _FASIL_PROMPT_BASE.format(kurallar_blok="")
    _PROMPT_CACHE['fasil'] = result
    return result


def build_pozisyon_prompt(conn):
    if 'pozisyon' in _PROMPT_CACHE:
        return _PROMPT_CACHE['pozisyon']
    blok = ""  # kurallar_blok Adım 1'den kaldırıldı — POZISYON SECIMI KURALLARI 1-5 yeterli
    result = _POZISYON_PROMPT_BASE.format(kurallar_blok=blok)
    _PROMPT_CACHE['pozisyon'] = result
    return result


def search_gtip_fts(conn, query, limit=20):
    c = conn.cursor()
    try:
        rows = c.execute("""
            SELECT gtip_code, tanim, tanim_hiyerarsi
            FROM gtip_fts WHERE gtip_fts MATCH ?
            ORDER BY rank
            LIMIT ?
        """, (query, limit)).fetchall()
        return rows
    except Exception:
        return []


def _product_search_words(title, desc, keywords, product_details, max_words=20):
    text = f"{title} {desc} {keywords} {product_details}".lower()
    words = sorted(
        set(re.findall(r'[a-zA-ZğüşıöçĞÜŞİÖÇ]{3,}', text)),
        key=len,
        reverse=True,
    )
    return [w for w in words if w.lower() not in _TEMU_STOP][:max_words]


def retrieve_ranked_gtips(conn, title, desc, keywords, product_details, top_n=50, per_query=14, filter_fasils=None):
    """
    Urun metninden kelimeler -> FTS; skorla birlestir. Cetvelde gercek satirlari getirir
    (sadece fasil basi sirali liste yerine ilgili 392x/732x vb. satirlari modele sunar).
    filter_fasils: set of int fasil_no — sadece bu fasillara ait GTIPler skorlara dahil edilir.
    """
    words = _product_search_words(title, desc, keywords, product_details, max_words=22)
    scores = {}
    for w in words:
        rows = search_gtip_fts(conn, w, limit=per_query)
        for idx, r in enumerate(rows):
            code = r[0]
            if filter_fasils is not None:
                d = re.sub(r'[^0-9]', '', code)
                if len(d) < 2:
                    continue
                if int(d[:2]) not in filter_fasils:
                    continue
            bump = max(1, per_query - idx)
            scores[code] = scores.get(code, 0) + bump
    if not scores:
        return []
    ordered = sorted(scores.keys(), key=lambda c: (-scores[c], c))[:top_n]
    cur = conn.cursor()
    out = []
    for code in ordered:
        row = cur.execute(
            "SELECT gtip_code, tanim, tanim_hiyerarsi FROM gtip WHERE gtip_code = ?",
            (code,),
        ).fetchone()
        if row:
            out.append(row)
    return out


def _json_from_balanced_braces(s):
    if not s:
        return None
    start = s.find('{')
    if start < 0:
        return None
    depth = 0
    in_str = False
    esc = False
    quote_ch = ''
    for i in range(start, len(s)):
        c = s[i]
        if in_str:
            if esc:
                esc = False
            elif c == '\\':
                esc = True
            elif c == quote_ch:
                in_str = False
            continue
        if c in '"\'':
            in_str = True
            quote_ch = c
            continue
        if c == '{':
            depth += 1
        elif c == '}':
            depth -= 1
            if depth == 0:
                try:
                    return json.loads(s[start : i + 1])
                except json.JSONDecodeError:
                    return None
    return None


def _count_tok(client, model, text):
    """Tek bir string için token sayısı. count_tokens API kullanır."""
    if not text:
        return 0
    try:
        r = client.beta.messages.count_tokens(
            model=model,
            messages=[{"role": "user", "content": text}],
            betas=["token-counting-2024-11-01", "prompt-caching-2024-07-31"],
        )
        return r.input_tokens
    except Exception:
        return None


def _fuzzy_get_list(parsed, canonical_key, aliases=()):
    """
    JSON dict'te canonical_key veya alias'lardan birini ara.
    Hiçbiri yoksa herhangi bir integer listesi döner.
    """
    for key in (canonical_key,) + aliases:
        if key in parsed:
            return parsed[key]
    # Fuzzy: key içinde canonical'ın ilk 4 harfi geçiyorsa kabul et
    prefix = canonical_key[:4].lower()
    for k, v in parsed.items():
        if prefix in k.lower() and isinstance(v, list):
            return v
    # Son çare: JSON'daki ilk integer listesi
    return next(
        (v for v in parsed.values() if isinstance(v, list)
         and v and all(isinstance(x, (int, float)) for x in v)),
        None
    )


def extract_first_json_object(text):
    """
    Claude yaniti: ```json ... ``` bloklari veya metindeki ilk dengeli JSON nesnesi.
    """
    if not text:
        return None
    s = text.strip()
    for block in re.findall(r'```(?:json)?\s*(.*?)```', s, re.DOTALL | re.IGNORECASE):
        got = _json_from_balanced_braces(block.strip())
        if got is not None:
            return got
    return _json_from_balanced_braces(s)


_GTIP_RE = re.compile(r'^\d{4}\.\d{2}\.\d{2}\.\d{2}\.\d{2}$')


def normalize_gtip_code(code):
    """Turk cetveli formati: XX.XX.XX.XX.XX"""
    if not code:
        return None
    s = str(code).strip().replace(' ', '').replace(',', '.')
    if not _GTIP_RE.match(s):
        return None
    return s


def gtip_exists(conn, code):
    if not code:
        return False
    row = conn.execute('SELECT 1 FROM gtip WHERE gtip_code = ?', (code,)).fetchone()
    return row is not None


def sanitize_classification(conn, result):
    """
    Claude bazen HS kalibinda ama TR cetvelinde olmayan kodlar uydurur (or. 3926.90.99.00.00).
    Sadece SQLite gtip tablosunda gercekten var olan kodlari birakir.
    """
    if not isinstance(result, dict):
        return result
    out = dict(result)

    raw_main = out.get('gtip_code', '') or ''
    norm = normalize_gtip_code(raw_main)
    if norm and gtip_exists(conn, norm):
        out['gtip_code'] = norm
    else:
        if raw_main:
            warn = f" [Model onerdigi kod veritabaninda yok: {raw_main}]"
            out['gerekce'] = (out.get('gerekce', '') + warn)[:2500]
            out['guven'] = 'dusuk'
        out['gtip_code'] = ''

    alts = out.get('alternatifler', [])
    if isinstance(alts, list):
        valid = []
        seen = set()
        main = out.get('gtip_code') or ''
        for a in alts:
            n = normalize_gtip_code(a)
            if n and gtip_exists(conn, n) and n not in seen and n != main:
                seen.add(n)
                valid.append(n)
        out['alternatifler'] = valid
    return out


# ---------------------------------------------------------------------------
# Claude API classification
# ---------------------------------------------------------------------------

SYSTEM_PROMPT = """Sen deneyimli bir Turk Gumruk Tarife siniflandirma uzmanisin. Girdi: urun tanimi +
TARIFE CETVELI VERILERI (fasil notlari ve GTIP satirlari). Cikti: tek bir 12 haneli kod + gerekce.

GENEL MUHAKEME (urun adi ezberleme yok; her kalemi basligin yasal tanimina gore ele):

1) ASIL FONKSIYON: Urun ne yapar? (mekanik tutma, olcu, yazi, giyim, elektriksel iletkenlik/yalitim,
   gida ile temas, insaat/yapi, oyuncak, vb.) Süs/marka/SEO metni fonksiyonu degistirmez.

2) SPESIFIKLIK: Ayni fasil icinde veya komsu basliklarda daha ozel bir alt pozisyon var mi?
   Varsa onu tercih et; "Digerleri"yi son care olarak kullan.

3) ELEME (kisa mantik): Yanlis sinif secimlerini cetveldeki BASLIK TANIMINA gore ele:
   - Ayni fasilda yan yana duran alt pozisyonlarin yasal kapsamlari farkli olabilir; hangi kodun
     gectigini FASIL NOTU ve listedeki kod/hiyerarsi tanimlarindan cikar. Daha "dar" veya "ozel"
     gorunen bir kodu, metindeki tanim urunle uyusmuyorsa secme; bir alt basligi baska birinin
     yerine koyma.
   - Pazarlama, magaza kategorisi veya SEO etiketi gumruk basliginin yasal anlamini degistirmez.
   - Teknik fonksiyon gerektiren basliklar: urun o fonksiyonu gercekten saglamiyorsa o basligi ele.
   - Baska fasilda cetvel metnine gore daha uygun ozel baslik varsa genel "diger"den once onu dusun.

4) FASIL NOTLARI VE CETVEL METNI: Birincil kaynak bu mesajdaki fasil notlari ile GTIP satir
   tanimlaridir; acik dahil/haric ifadelerini aynen uygula.

5) GEREKCE (Turkce, 3-5 cumle, madde isareti kullanabilirsin):
   - Secilen pozisyonun urunun ana fonksiyonuyla uyumu.
   - En az bir mantikli alternatif baslik veya fasil (genelde komsu veya sik karisan) ve NIYE
     uygun olmadigi: baslik tanimindaki zorunlu ozellik (or. yalitim, gida temasi, insaat malzemesi)
     urunda yoksa veya dar kapsam disinda kaliyorsa soyle. Urun metninde gecmeyen kelime,
     ornek urun veya kategori UYDURMA (ornek: baska bir urun adi yazmak yasak).

6) KOD KAYNAGI: gtip_code ve alternatifler SADECE bu mesajdaki TARIFE CETVELI listesindeki
   satirlardan birebir kopya olmali. Listede yoksa uydurma; en yakin gercek satiri sec veya bos birak.

7) ONCELIKLI GTIP (varsa): "METNE GORE ONCELIKLI" bolumu yalnizca metin-kelime eslesmesiyle
   siralanmistir; tek basina yeterli degildir. Nihai kod mutlaka fasil notu ve satir tanimiyla
   uyumlu olmali; celisirse cetvel metnini esas al.

Yanitini SADECE su JSON formatinda ver (gerekce ONCE, kod EN SONA):
{
  "gerekce": "Turkce muhakeme metni — once fonksiyon, sonra spesifiklik, sonra eleme",
  "guven": "yuksek|orta|dusuk",
  "alternatifler": ["YYYY.YY.YY.YY.YY"],
  "fasil": 39,
  "gtip_code": "XXXX.XX.XX.XX.XX"
}"""

REFINE_SYSTEM_PROMPT = """Ayni gorev: Turk gumruk GTIP siniflandirmasi. Onceki JSON cevabi zayif veya eksik olabilir.
TARIFE metnini tekrar dikkatle uygula; gtip_code ve alternatifler SADECE mesajdaki listede var olan
12 haneli kodlardan secilsin. Yanit SADECE gecerli JSON (gtip_code, fasil, gerekce, guven, alternatifler)."""

_BOLUM_PROMPT_BASE = """Sen deneyimli bir Turk Gumruk Tarife siniflandirma uzmanisin.
Gorev: Urun icin en uygun 5 aday BOLUMU belirle.
Sadece bolum listesine bak; fasil detayina girme.

KRITIK KURAL — FONKSIYON MATERYALI EZER:
Urunun KULLANIM AMACI ve FONKSIYONU her zaman yapildigi malzemenin onune gecer.
- "Tekstil kapli sac bandi" → sac aksesuari (Bolum 20), tekstil urunu degil (Bolum 11)
- "Kaucuk balik avlama boncugu" → balikcilik ekipmani (Bolum 20), kaucuk esyasi degil (Bolum 7)
- "Plastik tabanli metal aski" → metal esya (Bolum 15), plastik degil (Bolum 7)
Malzeme adi urun adinda gecse bile, urunun ne ISE YARADIGINA bak.

ONCELIK SIRASI: 1) Urun ne ise yarar? (fonksiyon)  2) Hangi sektorde kullanilir?  3) Malzeme

ISTISNA — TEKSTIL FORMU: Monofilament, elyaf, ip, halat/sicim/kordon formundaki urunler
kullanim amacinden bagimsiz olarak Bolum 11 (Dokumaya Elverişli Maddeler)'e girebilir.
Orn: balikcilik misinasi (monofilament → Bolum 11, Fasil 54), PE orgu ip (halat → Bolum 11, Fasil 56).

KRITIK: JSON key tam olarak "aday_bolumler" olmali, baska hicbir yazim kabul edilmez.
Yanitini SADECE asagidaki JSON formatinda ver (gerekce ONCE, sonra bolumler):
{"gerekce": "Once fonksiyon: urun ne ise yarar? Sonra bolum secimi (1-2 cumle)", "aday_bolumler": [20, 7, 11, 15, 9]}"""

_FASIL_PROMPT_BASE = """Sen deneyimli bir Turk Gumruk Tarife siniflandirma uzmanisin.
Gorev: Asagidaki fasil listesinden urun icin 5 aday FASIL belirle.

{kurallar_blok}

Yanitini SADECE su JSON formatinda ver (gerekce ONCE, sonra fasiller):
{{
  "gerekce": "Once fonksiyon: urun ne ise yarar? Sonra fasil secimi (2-3 cumle)",
  "aday_fasiller": [96, 33, 71, 39, 44]
}}"""

_POZISYON_PROMPT_BASE = """Sen deneyimli bir Turk Gumruk Tarife siniflandirma uzmanisin.
Gorev: Urun icin dogru FASIL ve 4 haneli POZİSYONU belirle.

Verilen fasil notlari, izahname ve pozisyon listesine gore en uygun pozisyonu sec.

{kurallar_blok}

- Fasil notu ve izahname dahil/haric hukumlerini aynen uygula.
- En ozel pozisyonu sec; "Digerleri"ni son care olarak kullan.
- Listede olmayan pozisyon uydurma.

POZISYON SECIMI KURALLARI:
1. MONTAJ YONTEMI SINIF DEGILDIR: Urunun montaj bicimi (kendinden yapiskanlı, vidalı,
   manyetik, kelepce) o urunun pozisyonunu belirlemez. Urunu asil fonksiyonu ve kullanim
   amaci belirler. Yapiskanlı plastik klips bir "klips"tir, yapiskanlı film/bant degildir.

2. HAM MALZEME FORMU vs MAMUL URUN: "Kendinden yapiskanlı levhalar, plakalar, bantlar,
   seritler, filmler, folyolar" gibi ifadeler, o formda HAM MALZEME olarak satilan urunler
   icin gecerlidir (orn. yapiskanlı rulo film, ambalaj banti). Belirli bir is icin tasarlanmis
   mamul plastik urunler (klips, tutucu, fitil, aksesuar, yapi elemani) bu pozisyona girmez;
   kendi amac/fonksiyon pozisyonlarinda siniflandirilir.

3. "DIGERLERI" MUTLAK SON CAREDIR: "Diger esya" veya "digerleri" iceren pozisyon YALNIZCA
   hicbir spesifik pozisyon uymadığinda kullanilir. Bir spesifik pozisyon ürünün amacini
   veya fonksiyonunu kapsiyorsa, o spesifik pozisyona git.
   Spesifik pozisyon "dar kapsami" gerekcesiyle reddedilip "digerleri"ne kacilmaz.

4. VE ILE BAGLANAN KAPSAMLAR: "sofra, mutfak, diger ev esyasi VE saglik/tuvalet esyasi"
   gibi birden fazla kategori iceren tanim, bunlardan HERHANGI birini kapsayan urunlere
   uygulanir. Banyo aksesuari = tuvalet esyasi kapsamindadir.

5. KURAL 3a — EN OZEL TANIM ONCELIKLIDIR: Birden fazla pozisyon aday oldugunda, esyayi
   EN OZEL sekilde tanimlayan pozisyon secilir. "Tarifenin baska pozisyonlarinda yer almayan"
   veya "baska yerde belirtilmemis" iceren pozisyon KALIFIKASYONU GEREGIDIR — ayni fasildaki
   veya aday fasillardaki baska bir pozisyon bu urunu kapsiyorsa, "baska yerde belirtilmemis"
   pozisyon gecersiz kalir. Once diger pozisyonlari kontrol et, hicbiri uymuyorsa "baska yerde
   belirtilmemis" pozisyonu sec.

Yanitini SADECE su JSON formatinda ver (degerlendirme ONCE, karar SONRA):
{{
  "degerlendirme": {{
    "8473": "Uyar: bilgisayar parcalari ve aksesuvarlari kapsaminda — urun bu tanima giriyor",
    "8471": "Uymaz: bilgisayarin kendisi, parca degil"
  }},
  "fasil": 84,
  "pozisyon_kod": "84.73"
}}"""

_POZISYON_1B_PROMPT = """Sen Turk Gumruk Tarife siniflandirma uzmanisin.
Gorev: Asagida verilen aday pozisyonlar arasından urun icin en dogru olanı sec.

Her pozisyon icin tanim ve izahname verilmistir.
Her pozisyon icin tanim ve izahname'yi birlikte oku; o pozisyonun tam kapsamini anla.
Urunun birincil kullanim amaci hangi pozisyonun kapsamina giriyorsa onu sec.

SADECE verilen aday pozisyonlar arasından sec — listede olmayan pozisyon ekleme.

Yanitini SADECE su JSON formatinda ver (degerlendirme ONCE, karar SONRA):
{{
  "degerlendirme": {{
    "9603": "Uyar: kozmetik tatbik fircalari bu pozisyona girer",
    "9615": "Uymaz: sac taraklari ve tokalar, urun fırça"
  }},
  "fasil": 96,
  "pozisyon_kod": "96.03"
}}"""


def _needs_refine(cls):
    if cls.get('error') or cls.get('parse_hatasi'):
        return False
    if not cls.get('gtip_code'):
        return True
    g = (cls.get('guven') or '').lower()
    return g in ('dusuk', 'orta')


def build_tarife_context(
    conn,
    title,
    desc,
    keywords,
    product_details,
    note_max_chars,
    gtip_rows_per_fasil,
    retrieval_top_n,
):
    ranked = retrieve_ranked_gtips(
        conn, title, desc, keywords, product_details, top_n=retrieval_top_n
    )
    parts = []
    if ranked:
        rlines = "\n".join(
            f"  {g[0]}  {g[1]}" + (f"  [{g[2]}]" if g[2] else "")
            for g in ranked
        )
        parts.append(f"=== METNE GORE ONCELIKLI GTIP (FTS skor) ===\n{rlines}")

    candidate_fasils = get_candidate_fasils(conn, product_details, keywords, desc, title)
    for fno in candidate_fasils[:6]:
        gtips = get_fasil_gtip_list(conn, fno, limit=200)
        if not gtips:
            continue
        note = get_fasil_notu(conn, fno)
        excerpt = (note[:note_max_chars] if note else "(not yok)")

        gtip_lines = "\n".join(
            f"  {g[0]}  {g[1]}" + (f"  [{g[2]}]" if g[2] else "")
            for g in gtips[:gtip_rows_per_fasil]
        )
        parts.append(
            f"=== FASIL {fno} ===\n"
            f"FASIL NOTU:\n{excerpt}\n\n"
            f"GTIP KODLARI:\n{gtip_lines}"
        )

    return "\n\n".join(parts), candidate_fasils


def build_pozisyon_context(conn, candidate_fasils, title, desc, keywords,
                           product_details, note_max_chars, retrieval_top_n,
                           izahname_max_chars=1500, return_atoms=False):
    """Adım 1 context: fasıl notları + izahname özeti + tüm 4'lü pozisyonlar."""
    fasils_for_context = candidate_fasils[:5]
    parts = []
    atoms = {}

    for fno in fasils_for_context:
        pozlar = get_all_pozisyonlar(conn, fno)
        if not pozlar:
            continue
        note = get_fasil_notu(conn, fno)
        excerpt = note[:note_max_chars] if (note and note_max_chars > 0) else ""
        izahname = get_izahname(conn, fno, izahname_max_chars)
        poz_lines = "\n".join(f"  {p[0]}  {p[1]}" for p in pozlar)
        if excerpt:
            atoms[f'fasil_{fno}_notu'] = excerpt
        atoms[f'fasil_{fno}_pozisyonlar'] = poz_lines
        if izahname:
            atoms[f'fasil_{fno}_izahname'] = izahname
        block = f"=== FASIL {fno} ===\n"
        if excerpt:
            block += f"FASIL NOTU:\n{excerpt}\n\n"
        if izahname:
            block += f"IZAHNAME (ozet):\n{izahname}\n\n"
        block += f"4'LU POZISYONLAR:\n{poz_lines}"
        parts.append(block)

    context = "\n\n".join(parts)
    return (context, atoms) if return_atoms else context


def _get_ara_pozisyonlar(conn, fasil_no, poz4_clean):
    """4'lü pozisyon altındaki ara seviye (6,8,10 hane) pozisyonları döner: {kod_clean: tanim}"""
    c = conn.cursor()
    rows = c.execute("""
        SELECT kod_clean, tanim FROM pozisyon
        WHERE fasil = ?
          AND length(kod_clean) > 4
          AND length(kod_clean) < 12
          AND substr(kod_clean, 1, 4) = ?
        ORDER BY length(kod_clean), kod_clean
    """, (fasil_no, poz4_clean)).fetchall()
    return {r[0]: r[1] for r in rows}


def _format_gtip_grouped(gtips, ara_pozlar):
    """12'li GTİP listesini ara pozisyon başlıklarıyla gruplu formatlar."""
    if not ara_pozlar:
        return "\n".join(f"  {g[0]}  {g[1]}" for g in gtips)

    def best_header(gtip_clean):
        best = None
        for ara_clean in ara_pozlar:
            if gtip_clean.startswith(ara_clean):
                if best is None or len(ara_clean) > len(best):
                    best = ara_clean
        return best

    lines = []
    last_header = None
    for g in gtips:
        gtip_clean = re.sub(r'[^0-9]', '', g[0])
        header = best_header(gtip_clean)
        if header and header != last_header:
            lines.append(f"\n  [{header}]  {ara_pozlar[header]}")
            last_header = header
        lines.append(f"    {g[0]}  {g[1]}")

    return "\n".join(lines).lstrip("\n")


def _get_pozisyon_tanim(conn, fasil_no, poz4_clean):
    row = conn.execute("""
        SELECT tanim FROM pozisyon
        WHERE fasil = ? AND substr(kod_clean, 1, 4) = ?
        ORDER BY seviye LIMIT 1
    """, (fasil_no, poz4_clean)).fetchone()
    return row[0] if row else ""


def build_gtip_context(conn, fasil_no, pozisyon_kod, note_max_chars, izahname_max_chars,
                       return_atoms=False):
    """Adım 2 context: fasıl notu + seçilen pozisyonun tüm 12'lileri (ara seviye başlıklı)."""
    poz4_clean = re.sub(r'[^0-9]', '', str(pozisyon_kod))[:4]
    gtips = get_gtips_by_pozisyon(conn, pozisyon_kod)
    note = get_fasil_notu(conn, fasil_no)
    izahname = get_izahname(conn, fasil_no, izahname_max_chars)
    poz_tanim = _get_pozisyon_tanim(conn, fasil_no, poz4_clean)

    parts = []
    atoms = {}
    note_excerpt = (note[:note_max_chars] if note else None)
    if note_excerpt:
        atoms['fasil_notu'] = note_excerpt
        parts.append(f"=== FASIL {fasil_no} NOTU ===\n{note_excerpt}")
    if izahname:
        atoms['izahname'] = izahname
        parts.append(f"=== FASIL {fasil_no} IZAHNAME (ozet) ===\n{izahname}")
    if gtips:
        ara_pozlar = _get_ara_pozisyonlar(conn, fasil_no, poz4_clean)
        gtip_block = _format_gtip_grouped(gtips, ara_pozlar)
        atoms['gtip_listesi'] = gtip_block
        header = f"POZISYON {pozisyon_kod}"
        if poz_tanim:
            header += f" — {poz_tanim}"
        parts.append(f"=== {header} ALTINDAKI TUM GTIP'LER ===\n{gtip_block}")

    context = "\n\n".join(parts)
    if return_atoms:
        return context, gtips, atoms
    return context, gtips


def build_gtip_context_multi(conn, aday_pozisyonlar, note_max_chars, izahname_max_chars):
    """
    2 aday pozisyon için birleşik context:
      Her pozisyon için: tanımı başlıkta göster + tüm 12'li GTİP'ler
    Adım 2'ye hem tanım hem tam liste gider; model pozisyon seçimini + GTİP seçimini birlikte yapar.
    """
    parts = []
    all_gtips = []
    primary_fasil = None

    for poz_kod in aday_pozisyonlar:
        poz4_clean = re.sub(r'[^0-9]', '', str(poz_kod))[:4]
        fasil_no = int(poz4_clean[:2]) if poz4_clean.isdigit() else None
        if primary_fasil is None:
            primary_fasil = fasil_no

        gtips = get_gtips_by_pozisyon(conn, poz_kod)
        if not gtips:
            continue
        all_gtips.extend(gtips)

        poz_tanim = _get_pozisyon_tanim(conn, fasil_no, poz4_clean) if fasil_no else ""
        ara_pozlar = _get_ara_pozisyonlar(conn, fasil_no, poz4_clean) if fasil_no else {}
        gtip_block = _format_gtip_grouped(gtips, ara_pozlar)

        header = f"ADAY POZISYON {poz_kod}"
        if poz_tanim:
            header += f" — {poz_tanim}"
        parts.append(f"=== {header} ===\n{gtip_block}")

    # Fasıl notu sadece primary fasıl için, max_chars > 0 ise
    if primary_fasil and note_max_chars > 0:
        note = get_fasil_notu(conn, primary_fasil)
        if note:
            parts.insert(0, f"=== FASIL {primary_fasil} NOTU ===\n{note[:note_max_chars]}")

    return "\n\n".join(parts), all_gtips


def _call_classify(client, model, max_tokens, system_prompt, user_msg):
    """Prompt caching olmadan basit API çağrısı (kısa system prompt'lar için)."""
    return client.messages.create(
        model=model,
        max_tokens=max_tokens,
        temperature=0,
        system=system_prompt,
        messages=[{"role": "user", "content": user_msg}],
    )


def _call_classify_ctx(client, model, max_tokens, system_prompt, context_text, query_text):
    """
    Prompt caching ile API çağrısı: context_text ayrı bir content block olarak
    cache_control ile işaretlenir. Aynı context (aynı pozisyon/fasıl) ile yapılan
    sonraki çağrılarda cache_read_input_tokens artar.
    Minimum 2048 token altındaysa API cache_control'ü sessizce görmezden gelir.
    """
    return client.messages.create(
        model=model,
        max_tokens=max_tokens,
        temperature=0,
        system=system_prompt,
        messages=[{
            "role": "user",
            "content": [
                {"type": "text", "text": context_text,
                 "cache_control": {"type": "ephemeral"}},
                {"type": "text", "text": query_text},
            ]
        }],
        extra_headers={"anthropic-beta": "prompt-caching-2024-07-31"},
    )


def _api_call_with_retry(client, model, max_tokens, system_prompt, user_msg):
    """Rate limit retry mantığıyla API çağrısı yapar (kısa mesajlar için)."""
    try:
        return _call_classify(client, model, max_tokens, system_prompt, user_msg)
    except anthropic.RateLimitError:
        for wait in [30, 60]:
            print(f"\n    Rate limit, {wait}s bekleniyor...", end="", flush=True)
            time.sleep(wait)
            try:
                return _call_classify(client, model, max_tokens, system_prompt, user_msg)
            except anthropic.RateLimitError:
                continue
        raise


def _api_call_ctx_with_retry(client, model, max_tokens, system_prompt, context_text, query_text):
    """Rate limit retry mantığıyla context-cached API çağrısı."""
    try:
        return _call_classify_ctx(client, model, max_tokens, system_prompt, context_text, query_text)
    except anthropic.RateLimitError:
        for wait in [30, 60]:
            print(f"\n    Rate limit, {wait}s bekleniyor...", end="", flush=True)
            time.sleep(wait)
            try:
                return _call_classify_ctx(client, model, max_tokens, system_prompt, context_text, query_text)
            except anthropic.RateLimitError:
                continue
        raise


def classify_product(client, product_info, conn, opts=None):
    """
    2-adımlı hiyerarşik sınıflandırma:
      Adım 1: Aday fasıllar → 4'lü pozisyon seçimi
      Adım 2: Seçilen pozisyon altındaki tüm 12'liler → GTİP seçimi

    opts: model, max_tokens, note_max_chars, izahname_max_chars,
          gtip_rows_per_fasil (fallback için), retrieval_top_n,
          refine, refine_model, refine_max_tokens
    """
    opts = opts or {}
    model               = opts.get('model', 'claude-haiku-4-5-20251001')
    max_tokens          = int(opts.get('max_tokens', 1200))
    note_max_chars      = int(opts.get('note_max_chars', 0))
    izahname_max_chars  = int(opts.get('izahname_max_chars', 0))
    gtip_rows_per_fasil = int(opts.get('gtip_rows_per_fasil', 120))
    retrieval_top_n     = int(opts.get('retrieval_top_n', 50))
    do_refine           = bool(opts.get('refine'))
    refine_model        = opts.get('refine_model', 'claude-sonnet-4-20250514')
    refine_max_tokens   = int(opts.get('refine_max_tokens', 1200))
    do_adim1b           = bool(opts.get('adim1b', True))   # Adım 1b izahname doğrulama
    adim1b_model        = opts.get('adim1b_model', 'claude-sonnet-4-20250514')  # 1b sabit sonnet
    do_token_breakdown  = bool(opts.get('token_breakdown'))

    title           = product_info.get('title', '')
    desc            = product_info.get('description', '')
    keywords        = product_info.get('keywords', '')
    product_details = product_info.get('product_details', '')
    sku_variants    = product_info.get('sku_variants', '')

    product_text = (
        f"Baslik: {title}\n"
        f"Aciklama: {desc}\n"
        f"Urun Detaylari: {product_details or '(belirtilmemis)'}\n"
        f"Varyantlar: {sku_variants or '(yok)'}"
    )

    # ------------------------------------------------------------------
    # ADIM 0a — Bölüm seçimi (21 bölüm → 2 aday bölüm)
    # ------------------------------------------------------------------
    bolum_system_prompt    = build_bolum_prompt()
    fasil_system_prompt    = build_fasil_prompt(conn)
    pozisyon_system_prompt = build_pozisyon_prompt(conn)

    bolum_listesi = get_bolum_listesi(conn)
    bolum_text = "\n".join(f"  Bolum {b[0]:2d}: {b[1]}" for b in bolum_listesi)
    bolum_user_msg = (
        f"Asagidaki urun icin dogru bolumu sec.\n\n"
        f"URUN BILGILERI:\n{product_text}\n\n"
        f"TURK GUMRUK TARIFE BOLUMLERI:\n{bolum_text}\n\n"
        f"Yanitini SADECE JSON olarak ver."
    )

    # Token breakdown sözlüğü — adım adım atomlar
    tbd = {}  # token_breakdown dict

    if do_token_breakdown:
        ct = lambda text: _count_tok(client, model, text)
        tbd['adim_0a'] = {
            'system_prompt':    ct(bolum_system_prompt),
            'urun_baslik':      ct(f"Baslik: {title}"),
            'urun_aciklama':    ct(f"Aciklama: {desc}"),
            'urun_detaylar':    ct(f"Urun Detaylari: {product_details or '(belirtilmemis)'}"),
            'urun_variantlar':  ct(f"Varyantlar: {sku_variants or '(yok)'}"),
            'bolum_listesi':    ct(bolum_text),
        }

    candidate_bolumler = []
    bolum_raw_response = None
    usage_0a = None
    try:
        bolum_resp = _api_call_with_retry(client, model, 400, bolum_system_prompt, bolum_user_msg)
        bolum_raw_response = bolum_resp.content[0].text
        usage_0a = {'in': bolum_resp.usage.input_tokens, 'out': bolum_resp.usage.output_tokens,
                    'cache_write': getattr(bolum_resp.usage, 'cache_creation_input_tokens', 0) or 0,
                    'cache_read':  getattr(bolum_resp.usage, 'cache_read_input_tokens', 0) or 0}
        bolum_parsed = extract_first_json_object(bolum_raw_response)
        if bolum_parsed:
            raw = _fuzzy_get_list(bolum_parsed, 'aday_bolumler',
                                  aliases=('anad_bolumler', 'adat_bolumler',
                                           'candidate_bolumler', 'aday_bolum'))
            if raw:
                candidate_bolumler = [int(float(x)) for x in raw
                                      if isinstance(x, (int, float)) and 1 <= int(float(x)) <= 21][:5]
    except Exception:
        pass

    # ------------------------------------------------------------------
    # ADIM 0b — Fasıl seçimi (seçilen bölümlerin fasılları → 3 aday fasıl)
    # ------------------------------------------------------------------
    candidate_fasils = []
    fasil_raw_response = None
    fasil_user_msg = None
    usage_0b = None
    gtip_context_block = None
    gtip_query = None
    gtip_raw_response = None
    if candidate_bolumler:
        fasiller = get_fasiller_by_bolumler(conn, candidate_bolumler)
        fasil_text = "\n".join(f"  Fasil {f[0]:02d}: {f[1]}" for f in fasiller)
        if do_token_breakdown:
            tbd['adim_0b'] = {
                'system_prompt': ct(fasil_system_prompt),
                'urun_metni':    ct(product_text),
                'fasil_listesi': ct(fasil_text),
            }
        fasil_user_msg = (
            f"Asagidaki urun icin dogru fasillari sec.\n\n"
            f"URUN BILGILERI:\n{product_text}\n\n"
            f"FASIL LISTESI:\n{fasil_text}\n\n"
            f"Yanitini SADECE JSON olarak ver."
        )
        try:
            fasil_resp = _api_call_with_retry(client, model, 400, fasil_system_prompt, fasil_user_msg)
            fasil_raw_response = fasil_resp.content[0].text
            usage_0b = {'in': fasil_resp.usage.input_tokens, 'out': fasil_resp.usage.output_tokens,
                        'cache_write': getattr(fasil_resp.usage, 'cache_creation_input_tokens', 0) or 0,
                        'cache_read':  getattr(fasil_resp.usage, 'cache_read_input_tokens', 0) or 0}
            fasil_parsed = extract_first_json_object(fasil_raw_response)
            if fasil_parsed:
                raw = _fuzzy_get_list(fasil_parsed, 'aday_fasiller',
                                      aliases=('aday_fasil', 'candidate_fasiller',
                                               'anad_fasiller', 'adat_fasiller'))
                if raw:
                    candidate_fasils = [int(float(x)) for x in raw
                                        if isinstance(x, (int, float)) and 1 <= int(float(x)) <= 97][:8]
        except Exception:
            pass

    # Adım 0 başarısız olursa FTS fallback
    if not candidate_fasils:
        candidate_fasils = get_candidate_fasils(conn, product_details, keywords, desc, title)

    # ------------------------------------------------------------------
    # ADIM 1 — Pozisyon seçimi
    # ------------------------------------------------------------------
    poz_context, poz_atoms = build_pozisyon_context(
        conn, candidate_fasils, title, desc, keywords,
        product_details, note_max_chars, retrieval_top_n,
        izahname_max_chars=izahname_max_chars, return_atoms=True,
    )
    if do_token_breakdown:
        tbd['adim_1'] = {'system_prompt': ct(pozisyon_system_prompt), 'urun_metni': ct(product_text)}
        tbd['adim_1'].update({k: ct(v) for k, v in poz_atoms.items()})
    poz_context_block = f"TARIFE CETVELI:\n{poz_context}"
    poz_query = (
        f"Asagidaki urun icin dogru FASIL ve 4 haneli POZISYONU sec.\n\n"
        f"URUN BILGILERI:\n{product_text}\n\n"
        f"Yukaridaki tarife cetvelini kullan.\n\n"
        f"Yanitini SADECE JSON olarak ver."
    )

    poz_result = None
    pozisyon_raw_response = None
    usage_1 = None
    adim1b_raw_response = None
    usage_1b = None
    try:
        poz_resp = _api_call_ctx_with_retry(client, model, 900, pozisyon_system_prompt,
                                            poz_context_block, poz_query)
        pozisyon_raw_response = poz_resp.content[0].text
        usage_1 = {'in': poz_resp.usage.input_tokens, 'out': poz_resp.usage.output_tokens,
                   'cache_write': getattr(poz_resp.usage, 'cache_creation_input_tokens', 0) or 0,
                   'cache_read':  getattr(poz_resp.usage, 'cache_read_input_tokens', 0) or 0}
        poz_parsed = extract_first_json_object(pozisyon_raw_response)
        if poz_parsed and poz_parsed.get('pozisyon_kod'):
            poz_result = poz_parsed
    except Exception:
        pass

    # ------------------------------------------------------------------
    # ADIM 1b — İzahname doğrulaması
    # Adım 1a'nın degerlendirme dict'indeki pozisyonlar için izahname
    # bölümleri çekilir; model kararını gözden geçirir.
    # ------------------------------------------------------------------
    if do_adim1b and poz_result:
        deger_dict = poz_result.get('degerlendirme') or {}
        iz_bloklar = []
        for poz_key in deger_dict:
            poz4 = re.sub(r'[^0-9]', '', str(poz_key))[:4]
            if len(poz4) == 4:
                fasil_for_iz = int(poz4[:2])
                poz_tanim_1b = _get_pozisyon_tanim(conn, fasil_for_iz, poz4)
                snippet = get_izahname_for_pozisyon(conn, fasil_for_iz, poz4)
                blok = f"[{poz_key}]"
                if poz_tanim_1b:
                    blok += f" {poz_tanim_1b}"
                if snippet:
                    blok += f"\nİZAHNAME:\n{snippet}"
                if snippet or poz_tanim_1b:
                    iz_bloklar.append(blok)
        if iz_bloklar:
            iz_context = "\n\n---\n\n".join(iz_bloklar)
            adim1b_query = (
                f"URUN BILGILERI:\n{product_text}\n\n"
                f"Asagida aday pozisyonlarin tanim ve izahnameleri verildi.\n"
                f"Yanitini SADECE JSON olarak ver (degerlendirme ONCE, karar SONRA)."
            )
            try:
                resp1b = _api_call_ctx_with_retry(
                    client, adim1b_model, 1200, _POZISYON_1B_PROMPT,
                    f"ADAY POZİSYONLAR:\n{iz_context}", adim1b_query,
                )
                adim1b_raw_response = resp1b.content[0].text
                usage_1b = {
                    'in': resp1b.usage.input_tokens,
                    'out': resp1b.usage.output_tokens,
                    'cache_write': getattr(resp1b.usage, 'cache_creation_input_tokens', 0) or 0,
                    'cache_read':  getattr(resp1b.usage, 'cache_read_input_tokens', 0) or 0,
                }
                parsed_1b = extract_first_json_object(adim1b_raw_response)
                if parsed_1b and parsed_1b.get('pozisyon_kod'):
                    poz_result = parsed_1b
            except Exception:
                pass  # 1b başarısız olursa 1a sonucu korunur

    def _make_debug(pozisyon_kod=None, fasil_no=None, usage_2=None, tbd=tbd):
        token_log = {
            'adim_0a': usage_0a,
            'adim_0b': usage_0b,
            'adim_1':  usage_1,
            'adim_1b': usage_1b,
            'adim_2':  usage_2,
        }
        total_in    = sum(u['in']          for u in token_log.values() if u)
        total_out   = sum(u['out']         for u in token_log.values() if u)
        total_cw    = sum(u.get('cache_write', 0) for u in token_log.values() if u)
        total_cr    = sum(u.get('cache_read',  0) for u in token_log.values() if u)
        token_log['toplam'] = {'in': total_in, 'out': total_out,
                               'cache_write': total_cw, 'cache_read': total_cr}
        return {
            'candidate_bolumler':    candidate_bolumler,
            'bolum_system_prompt':   bolum_system_prompt,
            'bolum_user_msg':        bolum_user_msg,
            'bolum_raw_response':    bolum_raw_response,
            'candidate_fasiller':    candidate_fasils,
            'fasil_system_prompt':   fasil_system_prompt,
            'fasil_user_msg':        fasil_user_msg if candidate_bolumler else None,
            'fasil_raw_response':    fasil_raw_response,

            'fts_bloku':             poz_atoms.get('fts_bloku'),
            'pozisyon_system_prompt': pozisyon_system_prompt,
            'pozisyon_context_block': poz_context_block,
            'pozisyon_query':        poz_query,
            'pozisyon_raw_response': pozisyon_raw_response,
            'adim1b_raw_response':   adim1b_raw_response,
            'secilen_fasil':         fasil_no,
            'gtip_context_block':    gtip_context_block if pozisyon_kod else None,
            'gtip_query':            gtip_query if pozisyon_kod else None,
            'gtip_raw_response':     gtip_raw_response if pozisyon_kod else None,
            'token_usage':           token_log,
            'token_breakdown':       tbd if tbd else None,
        }

    # Pozisyon seçimi başarısızsa eski flat yönteme düş
    if not poz_result:
        out = _classify_flat(client, product_info, conn, opts,
                             candidate_fasils, note_max_chars,
                             gtip_rows_per_fasil, retrieval_top_n)
        dbg = _make_debug()
        dbg['flat_mode'] = 'adim_1_parse_fail'
        out['debug'] = dbg
        return out

    fasil_no     = poz_result.get('fasil')
    pozisyon_kod = str(poz_result.get('pozisyon_kod', '')).strip()

    gtips_check = get_gtips_by_pozisyon(conn, pozisyon_kod)
    if not gtips_check:
        out = _classify_flat(client, product_info, conn, opts,
                             candidate_fasils, note_max_chars,
                             gtip_rows_per_fasil, retrieval_top_n)
        dbg = _make_debug(pozisyon_kod, fasil_no)
        dbg['flat_mode'] = 'adim_1_pozisyon_db_yok'
        out['debug'] = dbg
        return out

    # ------------------------------------------------------------------
    # ADIM 2 — GTİP seçimi
    # ------------------------------------------------------------------
    kurallar = get_yorum_kurallari(conn)
    system_step2 = (
        f"TARIFEYE ILISKIN GENEL YORUMLAMA KURALLARI:\n{kurallar}\n\n---\n\n{SYSTEM_PROMPT}"
        if kurallar else SYSTEM_PROMPT
    )

    if do_token_breakdown:
        gtip_context, _, gtip_atoms = build_gtip_context(
            conn, fasil_no, pozisyon_kod, note_max_chars, izahname_max_chars,
            return_atoms=True,
        )
        tbd['adim_2'] = {'system_prompt': ct(system_step2), 'urun_metni': ct(product_text)}
        tbd['adim_2'].update({k: ct(v) for k, v in gtip_atoms.items()})
    else:
        gtip_context, _ = build_gtip_context(
            conn, fasil_no, pozisyon_kod, note_max_chars, izahname_max_chars,
        )

    gtip_context_block = f"TARIFE CETVELI VERILERI:\n{gtip_context}"
    gtip_query = (
        f"Asagidaki urun icin dogru 12 haneli GTIP kodunu belirle.\n\n"
        f"URUN BILGILERI:\n{product_text}\n"
        f"Secilen pozisyon: {pozisyon_kod} (Fasil {fasil_no})\n\n"
        f"Yukaridaki tarife cetvelini kullan.\n\n"
        f"Yanitini SADECE JSON olarak ver."
    )

    def run_step2(sys_p, mdl, mtok):
        resp = _api_call_ctx_with_retry(client, mdl, mtok, sys_p, gtip_context_block, gtip_query)
        usage = {'in': resp.usage.input_tokens, 'out': resp.usage.output_tokens,
                 'cache_write': getattr(resp.usage, 'cache_creation_input_tokens', 0) or 0,
                 'cache_read':  getattr(resp.usage, 'cache_read_input_tokens', 0) or 0}
        text = resp.content[0].text.strip()
        parsed = extract_first_json_object(text)
        if parsed is None:
            result = {"gtip_code": "", "gerekce": text[:300], "guven": "dusuk",
                      "error": "JSON parse edilemedi", "parse_hatasi": True}
        else:
            result = sanitize_classification(conn, parsed)
        result['_usage'] = usage
        result['_raw'] = text
        return result

    try:
        out = run_step2(system_step2, model, max_tokens)
        usage_2 = out.pop('_usage', None)
        gtip_raw_response = out.pop('_raw', None)
        out.pop('parse_hatasi', None)
        debug = _make_debug(pozisyon_kod, fasil_no, usage_2)
        out['debug'] = debug
        if do_refine and _needs_refine(out):
            refined = run_step2(REFINE_SYSTEM_PROMPT, refine_model, refine_max_tokens)
            refined.pop('parse_hatasi', None)
            if (not refined.get('error') and refined.get('gtip_code')
                    and gtip_exists(conn, refined['gtip_code'])):
                refined['gerekce'] = ('[Ikinci gecis] ' + str(refined.get('gerekce', '')))[:2500]
                refined['debug'] = debug
                return refined
        return out
    except Exception as e:
        return {"gtip_code": "", "gerekce": "", "guven": "", "error": str(e)[:100], "debug": locals().get('debug', {})}


def _classify_flat(client, product_info, conn, opts, candidate_fasils,
                   note_max_chars, gtip_rows_per_fasil, retrieval_top_n):
    """Eski tek-adımlı flat sınıflandırma (fallback)."""
    model             = opts.get('model', 'claude-haiku-4-5-20251001')
    max_tokens        = int(opts.get('max_tokens', 1200))
    do_refine         = bool(opts.get('refine'))
    refine_model      = opts.get('refine_model', 'claude-sonnet-4-20250514')
    refine_max_tokens = int(opts.get('refine_max_tokens', 1200))

    title           = product_info.get('title', '')
    desc            = product_info.get('description', '')
    keywords        = product_info.get('keywords', '')
    product_details = product_info.get('product_details', '')
    sku_variants    = product_info.get('sku_variants', '')

    tarife_context, _ = build_tarife_context(
        conn, title, desc, keywords, product_details,
        note_max_chars, gtip_rows_per_fasil, retrieval_top_n
    )
    user_msg = (
        f"Asagidaki urun icin dogru 12 haneli GTIP kodunu belirle.\n\n"
        f"URUN BILGILERI:\nBaslik: {title}\nAciklama: {desc}\n"
        f"Urun Detaylari: {product_details or '(belirtilmemis)'}\n"
        f"Varyantlar: {sku_variants or '(yok)'}\n\n"
        f"TARIFE CETVELI VERILERI:\n{tarife_context}\n\n"
        f"Yanitini SADECE JSON olarak ver."
    )

    def run_once(sys_p, mdl, mtok):
        resp = _api_call_with_retry(client, mdl, mtok, sys_p, user_msg)
        text = resp.content[0].text.strip()
        parsed = extract_first_json_object(text)
        if parsed is None:
            return {"gtip_code": "", "gerekce": text[:300], "guven": "dusuk",
                    "error": "JSON parse edilemedi", "parse_hatasi": True}
        return sanitize_classification(conn, parsed)

    try:
        out = run_once(SYSTEM_PROMPT, model, max_tokens)
        out.pop('parse_hatasi', None)
        if do_refine and _needs_refine(out):
            refined = run_once(REFINE_SYSTEM_PROMPT, refine_model, refine_max_tokens)
            refined.pop('parse_hatasi', None)
            if (not refined.get('error') and refined.get('gtip_code')
                    and gtip_exists(conn, refined['gtip_code'])):
                refined['gerekce'] = ('[Ikinci gecis] ' + str(refined.get('gerekce', '')))[:2500]
                return refined
        return out
    except Exception as e:
        return {"gtip_code": "", "gerekce": "", "guven": "", "error": str(e)[:100]}


# ---------------------------------------------------------------------------
# Excel I/O
# ---------------------------------------------------------------------------

def _slug_header(name):
    """Excel baslik hucrelerini karsilastirma icin normalize et."""
    if name is None:
        return ''
    s = str(name).lower()
    # Turkish character transliteration
    tr_map = str.maketrans('üöşıığçÜÖŞIİĞÇ', 'uosiigcuosiigc')
    s = s.translate(tr_map)
    s = re.sub(r'[^a-z0-9]+', '_', s)
    return s.strip('_')


def normalize_product_row(row):
    """
    Scraper satiri veya manuel import (or. product_url, product_title, category_path,
    thumbnail, product_details) -> classify_product / HTML icin tek sema.
    """
    by_slug = {}
    for k, v in row.items():
        if k is None:
            continue
        sk = _slug_header(k if isinstance(k, str) else str(k))
        by_slug[sk] = v if v is not None else ''

    def pick(*slugs):
        for s in slugs:
            if s in by_slug and str(by_slug[s]).strip():
                return str(by_slug[s]).strip()
        return ''

    url = pick('url', 'product_url', 'product_link', 'link')
    title = pick('title', 'product_title', 'urun_basligi', 'urun_adi',
                 'turkce_ticari_tanim', 'ticari_tanim', 'turkce_ticar_itanim', 'mal_tanimi')
    description = pick('description', 'aciklama', 'desc', 'urun_aciklamasi')
    keywords = pick('keywords', 'keyword', 'category_path', 'category', 'kategori')
    product_details = pick('product_details', 'urun_detaylari')
    if not product_details:
        for sk, val in by_slug.items():
            if 'product' in sk and 'detail' in sk and str(val).strip():
                product_details = str(val).strip()
                break

    image_url = pick('image_url', 'thumbnail_url', 'img_url', 'kucuk_resim_url', 'resim_url')
    if not image_url:
        for sk, val in by_slug.items():
            if 'thumbnail' in sk and str(val).strip():
                image_url = str(val).strip()
                break
            if 'image' in sk and 'url' in sk and str(val).strip():
                image_url = str(val).strip()
                break

    sku_variants = pick('sku_variants', 'properties', 'variants', 'varyantlar')
    goods_id = pick('goods_id', 'goodsid', 'item_id')
    if not goods_id and url:
        m = re.search(r'goods_id=(\d+)', url) or re.search(r'-g-(\d+)\.html', url)
        if m:
            goods_id = m.group(1)

    err = row.get('error', '')
    if err is None:
        err = ''

    return {
        'url': url,
        'goods_id': goods_id,
        'title': title,
        'description': description,
        'keywords': keywords,
        'product_details': product_details,
        'image_url': image_url,
        'sku_variants': sku_variants,
        'error': str(err),
    }


def read_scraped_excel(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    products = []
    for r in range(2, ws.max_row + 1):
        row = {}
        for i, h in enumerate(headers):
            if h:
                row[h.lower().replace(' ', '_')] = ws.cell(r, i + 1).value or ''
        products.append(normalize_product_row(row))
    return products


def write_classified_excel(products, classifications, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "GTIP Siniflandirma"

    out_headers = ['URL', 'Goods ID', 'Title', 'Product Details',
                   'GTIP Kodu', 'Fasil', 'Gerekce', 'Guven', 'Alternatifler', 'Hata']
    ws.append(out_headers)

    bold = openpyxl.styles.Font(bold=True)
    for c in range(1, len(out_headers) + 1):
        ws.cell(1, c).font = bold

    for prod, cls in zip(products, classifications):
        alts = cls.get('alternatifler', [])
        alt_str = ', '.join(alts) if isinstance(alts, list) else str(alts)
        ws.append([
            prod.get('url', ''),
            prod.get('goods_id', ''),
            prod.get('title', ''),
            prod.get('product_details', ''),
            cls.get('gtip_code', ''),
            cls.get('fasil', ''),
            cls.get('gerekce', ''),
            cls.get('guven', ''),
            alt_str,
            cls.get('error', ''),
        ])

    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 60
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['G'].width = 60
    wb.save(output_path)


def _esc(text):
    return (str(text) or '').replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;').replace('"', '&quot;')


def write_classified_html(products, classifications, output_path):
    html_path = os.path.splitext(output_path)[0] + '.html'

    cards = []
    for i, (prod, cls) in enumerate(zip(products, classifications), 1):
        gtip = cls.get('gtip_code', '')
        guven = cls.get('guven', '')
        gerekce = cls.get('gerekce', '')
        alts = cls.get('alternatifler', [])
        error = cls.get('error', '')
        fasil = cls.get('fasil', '')

        guven_class = {'yuksek': 'high', 'orta': 'mid', 'dusuk': 'low'}.get(guven, 'low')

        details_html = ''
        pd = prod.get('product_details', '')
        if pd:
            rows = ''
            for item in str(pd).split('; '):
                parts = item.split(': ', 1)
                if len(parts) == 2:
                    rows += f'<tr><td class="pk">{_esc(parts[0])}</td><td>{_esc(parts[1])}</td></tr>'
            details_html = f'<table class="props">{rows}</table>'

        alt_html = ''
        if alts and isinstance(alts, list) and alts[0]:
            alt_html = '<div class="alts">Alt: ' + ', '.join(f'<code>{_esc(a)}</code>' for a in alts) + '</div>'

        img_url = prod.get('image_url', '')
        img_html = f'<img src="{_esc(img_url)}" loading="lazy">' if img_url else '<div class="no-img">No image</div>'

        error_html = f'<div class="error">{_esc(error)}</div>' if error else ''

        cards.append(f'''
    <div class="card">
      <div class="card-img">{img_html}</div>
      <div class="card-body">
        <div class="card-num">#{i}</div>
        <h2><a href="{_esc(prod.get('url', ''))}" target="_blank">{_esc(prod.get('title', '')) or 'Untitled'}</a></h2>
        <div class="gtip-box {guven_class}">
          <span class="gtip-code">{_esc(gtip) or '?'}</span>
          <span class="gtip-badge">{_esc(guven)}</span>
          <span class="gtip-fasil">Fasil {_esc(str(fasil))}</span>
        </div>
        <p class="gerekce">{_esc(gerekce)}</p>
        {alt_html}
        {details_html}
        {error_html}
      </div>
    </div>''')

    ok = sum(1 for c in classifications if not c.get('error'))
    high = sum(1 for c in classifications if c.get('guven') == 'yuksek')

    html = f'''<!DOCTYPE html>
<html lang="tr">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>GTIP Siniflandirma ({len(products)} urun)</title>
<style>
  * {{ margin: 0; padding: 0; box-sizing: border-box; }}
  body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; background: #f0f2f5; color: #1a1a1a; padding: 20px; }}
  .header {{ max-width: 1100px; margin: 0 auto 24px; }}
  .header h1 {{ font-size: 24px; font-weight: 700; }}
  .header .stats {{ color: #666; margin-top: 4px; font-size: 14px; }}
  .stats span {{ margin-right: 16px; }}
  .card {{ max-width: 1100px; margin: 0 auto 16px; background: #fff; border-radius: 10px; box-shadow: 0 1px 3px rgba(0,0,0,.08); display: flex; overflow: hidden; }}
  .card-img {{ width: 200px; min-height: 200px; flex-shrink: 0; background: #f7f7f7; display: flex; align-items: center; justify-content: center; }}
  .card-img img {{ width: 100%; height: 100%; object-fit: cover; }}
  .no-img {{ color: #ccc; font-size: 13px; }}
  .card-body {{ padding: 16px 20px; flex: 1; min-width: 0; }}
  .card-num {{ font-size: 12px; color: #999; margin-bottom: 4px; }}
  h2 {{ font-size: 15px; font-weight: 600; margin-bottom: 10px; line-height: 1.3; }}
  h2 a {{ color: #1a1a1a; text-decoration: none; }}
  h2 a:hover {{ color: #e67e00; }}
  .gtip-box {{ display: flex; align-items: center; gap: 10px; padding: 8px 12px; border-radius: 6px; margin-bottom: 10px; }}
  .gtip-box.high {{ background: #e8f5e9; border-left: 4px solid #2e7d32; }}
  .gtip-box.mid {{ background: #fff8e1; border-left: 4px solid #f9a825; }}
  .gtip-box.low {{ background: #fce4ec; border-left: 4px solid #c62828; }}
  .gtip-code {{ font-family: 'Consolas', monospace; font-size: 16px; font-weight: 700; }}
  .gtip-badge {{ font-size: 11px; padding: 2px 8px; border-radius: 10px; font-weight: 600; text-transform: uppercase; }}
  .high .gtip-badge {{ background: #2e7d32; color: #fff; }}
  .mid .gtip-badge {{ background: #f9a825; color: #fff; }}
  .low .gtip-badge {{ background: #c62828; color: #fff; }}
  .gtip-fasil {{ font-size: 12px; color: #666; }}
  .gerekce {{ font-size: 13px; color: #444; margin-bottom: 8px; line-height: 1.4; }}
  .alts {{ font-size: 12px; color: #666; margin-bottom: 8px; }}
  .alts code {{ background: #f0f0f0; padding: 1px 5px; border-radius: 3px; }}
  .props {{ font-size: 12px; border-collapse: collapse; margin-bottom: 8px; }}
  .props tr {{ border-bottom: 1px solid #f0f0f0; }}
  .props td {{ padding: 2px 10px 2px 0; }}
  .pk {{ font-weight: 600; color: #333; white-space: nowrap; }}
  .error {{ padding: 6px 10px; background: #fff0f0; color: #c00; border-radius: 4px; font-size: 13px; }}
  @media (max-width: 700px) {{
    .card {{ flex-direction: column; }}
    .card-img {{ width: 100%; height: 180px; }}
  }}
</style>
</head>
<body>
<div class="header">
  <h1>GTIP Siniflandirma Sonuclari</h1>
  <div class="stats">
    <span>{len(products)} urun</span>
    <span>{ok} siniflandirildi</span>
    <span>{high} yuksek guven</span>
  </div>
</div>
{"".join(cards)}
</body>
</html>'''

    with open(html_path, 'w', encoding='utf-8') as f:
        f.write(html)
    return html_path


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description='GTIP Siniflandirici (Claude API)')
    parser.add_argument('input', help='Scraper ciktisi Excel dosyasi')
    parser.add_argument('-o', '--output', help='Cikti dosyasi')
    parser.add_argument('--db', default='data/gtip_2026.db', help='GTIP veritabani yolu')
    parser.add_argument('--delay', type=float, default=0.5, help='API istekleri arasi bekleme (saniye)')
    parser.add_argument(
        '--model',
        default='claude-haiku-4-5-20251001',
        help='Ilk gecis Claude model id',
    )
    parser.add_argument('--max-tokens', type=int, default=1200, help='Ilk gecis max_tokens')
    parser.add_argument(
        '--note-chars',
        type=int,
        default=0,
        metavar='N',
        help='Fasil notundan modele giden max karakter (0=kapali)',
    )
    parser.add_argument(
        '--gtip-rows',
        type=int,
        default=120,
        metavar='N',
        help='Fallback modda her aday fasil icin GTIP satir sayisi',
    )
    parser.add_argument(
        '--izahname-chars',
        type=int,
        default=0,
        metavar='N',
        help='Izahname metninden modele giden max karakter (0=kapali)',
    )
    parser.add_argument(
        '--retrieval',
        type=int,
        default=50,
        metavar='N',
        help='Urun metnine gore FTS ile getirilecek oncelikli GTIP satiri',
    )
    parser.add_argument(
        '--refine',
        action='store_true',
        help='guven dusuk/orta veya kod yoksa ikinci gecis (daha guclu model)',
    )
    parser.add_argument(
        '--refine-model',
        default='claude-sonnet-4-20250514',
        help='Ikinci gecis model id',
    )
    parser.add_argument('--refine-max-tokens', type=int, default=1200)
    parser.add_argument(
        '--no-adim1b',
        action='store_true',
        help='Adim 1b izahname dogrulama adimini atla',
    )
    parser.add_argument(
        '--adim1b-model',
        default=None,
        help='Adim 1b icin model (default: --model ile ayni)',
    )
    args = parser.parse_args()

    api_key = os.environ.get('ANTHROPIC_API_KEY', '')
    if not api_key:
        env_file = os.path.join(os.path.dirname(__file__), '..', '.env')
        if os.path.exists(env_file):
            with open(env_file) as f:
                for line in f:
                    if line.strip().startswith('ANTHROPIC_API_KEY='):
                        api_key = line.strip().split('=', 1)[1].strip().strip('"').strip("'")

    if not api_key:
        print("Hata: ANTHROPIC_API_KEY bulunamadi.")
        print("  .env dosyasina ANTHROPIC_API_KEY=sk-ant-... yazin")
        sys.exit(1)

    if not os.path.isfile(args.input):
        print(f"Hata: {args.input} bulunamadi")
        sys.exit(1)

    if not os.path.isfile(args.db):
        print(f"Hata: {args.db} bulunamadi. Once build_db.py calistirin.")
        sys.exit(1)

    output_path = args.output
    if not output_path:
        base = os.path.splitext(os.path.basename(args.input))[0]
        output_path = os.path.join('output', f'{base}_classified.xlsx')
    os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)

    conn = sqlite3.connect(args.db)
    client = anthropic.Anthropic(api_key=api_key)

    classify_opts = {
        'model': args.model,
        'max_tokens': args.max_tokens,
        'note_max_chars': args.note_chars,
        'izahname_max_chars': args.izahname_chars,
        'gtip_rows_per_fasil': args.gtip_rows,
        'retrieval_top_n': args.retrieval,
        'refine': args.refine,
        'refine_model': args.refine_model,
        'refine_max_tokens': args.refine_max_tokens,
        'adim1b': not args.no_adim1b,
        'adim1b_model': args.adim1b_model or 'claude-sonnet-4-20250514',
    }

    products = read_scraped_excel(args.input)
    print(f"Toplam urun: {len(products)}")

    if not products:
        print("Hata: Excel'de urun bulunamadi")
        sys.exit(1)

    classifications = []
    errors = 0

    for i, prod in enumerate(products, 1):
        title = prod.get('title', '')[:50]
        print(f"  [{i}/{len(products)}] {title}...", end=" ", flush=True)

        cls = classify_product(client, prod, conn, classify_opts)

        if cls.get('error'):
            print(f"HATA: {cls['error'][:60]}")
            errors += 1
        else:
            code = cls.get('gtip_code', '?')
            guven = cls.get('guven', '?')
            print(f"-> {code} ({guven})")

        classifications.append(cls)
        if i < len(products):
            time.sleep(args.delay)

    write_classified_excel(products, classifications, output_path)
    html_path = write_classified_html(products, classifications, output_path)
    conn.close()

    high = sum(1 for c in classifications if c.get('guven') == 'yuksek')
    print(f"\nToplam         : {len(products)}")
    print(f"Siniflandirildi: {len(products) - errors}")
    print(f"Yuksek guven   : {high}")
    print(f"Hata           : {errors}")
    print(f"Excel          : {output_path}")
    print(f"HTML           : {html_path}")


if __name__ == "__main__":
    main()
